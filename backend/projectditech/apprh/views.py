from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from django.contrib.auth import authenticate
from django.db import models
from .models import User, Employee, Service, EmployeeHistory, JobOffer, Candidate, Interview, LeaveRequest, LeaveBalance, Attendance, Contract, Payslip, PayslipBonus, PayslipDeduction, PaymentHistory, Document, PresenceTracking, TrainingPlan, Training, TrainingSession, Evaluation
from .serializers import (
    UserSerializer, EmployeeSerializer, ServiceSerializer,
    EmployeeHistorySerializer, LoginSerializer, JobOfferSerializer, CandidateSerializer, InterviewSerializer,
    LeaveRequestSerializer, LeaveBalanceSerializer, AttendanceSerializer, ContractSerializer, PayslipSerializer,
    PayslipBonusSerializer, PayslipDeductionSerializer, PaymentHistorySerializer,
    DocumentSerializer, CustomTokenObtainPairSerializer, PresenceTrackingSerializer,
    TrainingPlanSerializer, TrainingSerializer, TrainingSessionSerializer, EvaluationSerializer
)
from .models import EmployeeHistory
from datetime import date, timedelta
from django.utils import timezone
from io import BytesIO
import os
from django.http import FileResponse
from django.conf import settings
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


class CustomTokenObtainPairView(TokenObtainPairView):
    """Vue personnalisée pour l'obtention de token avec informations utilisateur"""
    serializer_class = CustomTokenObtainPairSerializer


@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        username = serializer.validated_data['username']
        password = serializer.validated_data['password']
        
        # Vérifier d'abord si l'utilisateur existe
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return Response({
                'error': 'Aucun compte n\'a été trouvé avec ce nom d\'utilisateur.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Vérifier si l'utilisateur est actif
        if not user.is_active:
            return Response({
                'error': 'Ce compte est désactivé. Contactez l\'administrateur.'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Authentifier l'utilisateur
        user = authenticate(username=username, password=password)
        
        if user:
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'user': UserSerializer(user).data
            })
        else:
            return Response({
                'error': 'Mot de passe incorrect. Veuillez réessayer.'
            }, status=status.HTTP_401_UNAUTHORIZED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    """Get comprehensive dashboard statistics - Section G: Tableaux de bord RH"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        from django.db.models import Q, Sum, Avg, Count, Max, Min
        from datetime import datetime, timedelta
        from calendar import monthrange
        
        # Obtenir la date actuelle
        today = timezone.now().date()
        current_month = today.month
        current_year = today.year
        now = timezone.now()
        
        # ========== 1. EFFECTIF TOTAL ==========
        total_staff = Employee.objects.filter(is_active=True).count()
    
        # ========== 2. RENOUVELLEMENT (ROTATION DU PERSONNEL) ==========
        new_employees_this_month = Employee.objects.filter(
            date_of_hire__year=current_year,
            date_of_hire__month=current_month
        ).count()
        
        # Rotation du personnel (embauches - départs ce mois)
        departures_this_month = Employee.objects.filter(
            is_active=False,
            updated_at__year=current_year,
            updated_at__month=current_month
        ).count()
        staff_rotation = new_employees_this_month - departures_this_month
        
        # Effectif par service
        staff_by_service = Service.objects.annotate(
            employee_count=Count('employees', filter=Q(employees__is_active=True))
        ).values('id', 'name', 'employee_count')
        
        # ========== CONTRATS ==========
        total_contracts = Contract.objects.filter(status__in=['PENDING', 'SIGNED']).count()
        contracts_to_renew = Contract.objects.filter(needs_renewal=True).count()
        
        # Alertes contrats expirants
        ninety_days_later = today + timedelta(days=90)
        expiring_contracts = Contract.objects.filter(
            status='SIGNED',
            end_date__lte=ninety_days_later,
            end_date__gte=today
        ).select_related('employee', 'employee__service').order_by('end_date')
        
        contract_alerts = [
            {
                'id': contract.id,
                'employee_name': contract.employee.get_full_name(),
                'employee_id': contract.employee.employee_id,
                'service_name': contract.employee.service.name if contract.employee.service else 'Non assigné',
                'contract_type': contract.get_contract_type_display(),
                'end_date': contract.end_date.strftime('%Y-%m-%d'),
                'days_left': (contract.end_date - today).days,
                'alert_level': contract.alert_level,
                'needs_renewal': contract.needs_renewal
            }
            for contract in expiring_contracts
        ]
        
        # ========== PRÉSENCE & ABSENCES ==========
        present_today_count = PresenceTracking.objects.filter(
            date=today,
            status__in=['PRESENT', 'LATE']
        ).count()
        
        absent_today_count = PresenceTracking.objects.filter(
            date=today,
            status='ABSENT'
        ).count()
        
        late_today_count = PresenceTracking.objects.filter(
            date=today,
            is_late=True
        ).count()
        
        # Congés en cours (employés en congé aujourd'hui)
        employees_on_leave_today = LeaveRequest.objects.filter(
            Q(status='MANAGER_APPROVED') | Q(status='RH_APPROVED'),
            start_date__lte=today,
            end_date__gte=today
        ).values_list('employee_id', flat=True).distinct().count()
        
        current_leaves = employees_on_leave_today
        
        # Taux de présence : présents / (effectif total - en congé)
        staff_available_today = total_staff - employees_on_leave_today
        presence_rate = (present_today_count / staff_available_today * 100) if staff_available_today > 0 else 0.0
        
        # Demandes de congés en attente
        pending_leaves = LeaveRequest.objects.filter(status='PENDING').count()
        
        # Total absences/congés
        absences_leaves = LeaveRequest.objects.filter(
            Q(status='PENDING') | Q(status='MANAGER_APPROVED') | Q(status='RH_APPROVED')
        ).count()
        
        # ========== PAIE ==========
        # Masse salariale mensuelle
        monthly_payroll = Payslip.objects.filter(
            month=current_month,
            year=current_year
        ).aggregate(total=Sum('net_salary'))['total'] or 0
        
        # Masse salariale annuelle
        annual_payroll = Payslip.objects.filter(
            year=current_year
        ).aggregate(total=Sum('net_salary'))['total'] or 0
        
        # Salaire moyen
        avg_salary = Employee.objects.filter(is_active=True).aggregate(
            avg=Avg('salary')
        )['avg'] or 0
        
        # ========== RECRUTEMENT ==========
        active_job_offers = JobOffer.objects.filter(status='PUBLISHED').count()
        total_candidates = Candidate.objects.count()
        new_candidates_this_month = Candidate.objects.filter(
            application_date__year=current_year,
            application_date__month=current_month
        ).count()
        candidates_in_process = Candidate.objects.filter(
            status__in=['SCREENING', 'INTERVIEW', 'SECOND_INTERVIEW', 'FINAL_REVIEW']
        ).count()
        upcoming_interviews = Interview.objects.filter(
            scheduled_date__gte=now,
            status__in=['SCHEDULED', 'RESCHEDULED']
        ).count()
        
        # ========== FORMATION ==========
        active_training_plans = TrainingPlan.objects.filter(status='ACTIVE').count()
        trainings_in_progress = Training.objects.filter(status='IN_PROGRESS').count()
        completed_trainings_this_month = Training.objects.filter(
            status='COMPLETED',
            end_date__year=current_year,
            end_date__month=current_month
        ).count()
        
        # ========== ÉVALUATIONS ==========
        pending_evaluations = Evaluation.objects.filter(
            status__in=['DRAFT', 'IN_PROGRESS']
        ).count()
        annual_evaluations_this_year = Evaluation.objects.filter(
            evaluation_type='ANNUAL',
            evaluation_date__year=current_year
        ).count()
        
        # ========== DONNÉES HISTORIQUES (12 DERNIERS MOIS) ==========
        monthly_payroll_history = []
        presence_rate_history = []
        staff_count_history = []
        leaves_count_history = []
        months_labels = []
        
        for i in range(11, -1, -1):  # 12 mois en arrière
            target_date = now - timedelta(days=30 * i)
            target_month = target_date.month
            target_year = target_date.year
            
            months_labels.append(f"{target_date.strftime('%b %Y')}")
            
            # Masse salariale mensuelle réelle
            month_payroll = Payslip.objects.filter(
                month=target_month,
                year=target_year
            ).aggregate(total=Sum('net_salary'))['total'] or 0
            monthly_payroll_history.append(float(month_payroll))
            
            # Effectif à la fin du mois
            month_end = datetime(target_year, target_month, monthrange(target_year, target_month)[1]).date()
            staff_count = Employee.objects.filter(
                is_active=True,
                date_of_hire__lte=month_end
            ).count()
            staff_count_history.append(staff_count)
            
            # Taux de présence moyen du mois
            month_start = datetime(target_year, target_month, 1).date()
            # Compter les jours ouvrés du mois
            working_days = 0
            current_check = month_start
            while current_check <= month_end:
                # 0 = lundi, 6 = dimanche
                if current_check.weekday() < 5:  # Du lundi au vendredi
                    working_days += 1
                current_check += timedelta(days=1)
            
            # Compter les pointages présents
            month_presence = PresenceTracking.objects.filter(
                date__gte=month_start,
                date__lte=month_end,
                status__in=['PRESENT', 'LATE']
            ).count()
            
            # Calculer le taux de présence : pointages présents / (effectif * jours ouvrés)
            month_staff_count = Employee.objects.filter(
                is_active=True,
                date_of_hire__lte=month_end
            ).count()
            expected_presence = month_staff_count * working_days
            month_presence_rate = (month_presence / expected_presence * 100) if expected_presence > 0 else 0
            presence_rate_history.append(round(month_presence_rate, 1))
            
            # Nombre de congés du mois
            month_leaves = LeaveRequest.objects.filter(
                Q(status='MANAGER_APPROVED') | Q(status='RH_APPROVED'),
                start_date__lte=month_end,
                end_date__gte=month_start
            ).count()
            leaves_count_history.append(month_leaves)
        
        # ========== STATISTIQUES PAR SERVICE ==========
        service_stats = []
        for service in Service.objects.all():
            service_employees = Employee.objects.filter(service=service, is_active=True)
            service_count = service_employees.count()
            if service_count > 0:
                service_stats.append({
                    'service_id': service.id,
                    'service_name': service.name,
                    'employee_count': service_count,
                    'avg_salary': float(service_employees.aggregate(avg=Avg('salary'))['avg'] or 0),
                    'presence_today': PresenceTracking.objects.filter(
                        date=today,
                        employee__service=service,
                        status__in=['PRESENT', 'LATE']
                    ).count()
                })
        
        # ========== ALERTES ==========
        contract_alerts_count = len(contract_alerts)
        alerts = {
            'contracts_expiring': contract_alerts_count,
            'contracts_expiring_critical': len([a for a in contract_alerts if a.get('alert_level') == 'critical']),
            'pending_leaves': pending_leaves,
            'upcoming_interviews': upcoming_interviews,
            'pending_evaluations': pending_evaluations,
            'contracts_to_renew': contracts_to_renew
        }
        
        # ========== RÉSUMÉ - Section G: Tableaux de bord RH ==========
        # Structure selon les besoins spécifiques demandés
        stats = {
            # 1. Effectif total
            'effectif_total': total_staff,
            
            # 2. Renouvellement (Rotation du personnel)
            'renouvellement': {
                'embauches_ce_mois': new_employees_this_month,
                'departs_ce_mois': departures_this_month,
                'rotation': staff_rotation,
                'pourcentage_rotation': round((staff_rotation / total_staff * 100) if total_staff > 0 else 0, 2)
            },
            
            # 3. Absences/congés
            'absences_conges': {
                'total': absences_leaves,
                'conges_en_cours': current_leaves,
                'conges_en_attente': pending_leaves,
                'absences_aujourd_hui': absent_today_count
            },
            
            # 4. Masse Salariale Mensuelle
            'masse_salariale_mensuelle': float(monthly_payroll),
            'masse_salariale_annuelle': float(annual_payroll),
            'salaire_moyen': float(avg_salary),
            
            # 5. Taux de présence
            'taux_presence': round(presence_rate, 2),
            'presence_details': {
                'present': present_today_count,
                'absent': absent_today_count,
                'en_retard': late_today_count
            },
            
            # 6. Nombre de congés en cours
            'nombre_conges_en_cours': current_leaves,
            
            # 7. Alertes contrats expirants
            'alertes_contrats_expirants': {
                'total': len(contract_alerts),
                'critiques': len([a for a in contract_alerts if a.get('days_left', 999) <= 30]),
                'avertissements': len([a for a in contract_alerts if 30 < a.get('days_left', 999) <= 60]),
                'informations': len([a for a in contract_alerts if 60 < a.get('days_left', 999) <= 90]),
                'details': contract_alerts
            },
            
            # 8. Graphiques et statistiques
            'graphiques_statistiques': {
                'masse_salariale_mensuelle_historique': monthly_payroll_history,
                'taux_presence_historique': presence_rate_history,
                'effectif_historique': staff_count_history,
                'conges_historique': leaves_count_history,
                'labels_mois': months_labels,
                'effectif_par_service': list(staff_by_service),
                'statistiques_services': service_stats
            },
            
            # Données supplémentaires pour compatibilité
            'total_staff': total_staff,
            'total_employees': total_staff,
            'presence_rate': round(presence_rate, 1),
            'monthly_payroll': float(monthly_payroll),
            'current_leaves': current_leaves,
            'contract_alerts': contract_alerts,
            'contract_alerts_count': len(contract_alerts),
            'monthly_payroll_history': monthly_payroll_history,
            'presence_rate_history': presence_rate_history,
            'staff_count_history': staff_count_history,
            'leaves_count_history': leaves_count_history,
            'months_labels': months_labels
        }
        
        return Response(stats)
        
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des statistiques du dashboard: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return Response(
            {
                'error': 'Erreur lors de la récupération des statistiques',
                'detail': str(e) if settings.DEBUG else 'Une erreur est survenue. Contactez l\'administrateur.'
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_hr_analytics(request):
    """Analyses RH détaillées pour tableaux de bord"""
    from django.db.models import Q, Sum, Avg, Count
    from datetime import datetime, timedelta
    
    today = timezone.now().date()
    current_year = today.year
    
    # ========== ANALYSES DE PRÉSENCE ==========
    # Taux de présence mensuel moyen
    monthly_presence_avg = []
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1).date()
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1).date() - timedelta(days=1)
        
        total_days = PresenceTracking.objects.filter(
            date__gte=month_start,
            date__lte=month_end
        ).count()
        present_days = PresenceTracking.objects.filter(
            date__gte=month_start,
            date__lte=month_end,
            status__in=['PRESENT', 'LATE']
        ).count()
        
        rate = (present_days / total_days * 100) if total_days > 0 else 0
        monthly_presence_avg.append(round(rate, 1))
    
    # Retards par mois
    monthly_lates = []
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1).date()
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1).date() - timedelta(days=1)
        
        lates_count = PresenceTracking.objects.filter(
            date__gte=month_start,
            date__lte=month_end,
            is_late=True
        ).count()
        monthly_lates.append(lates_count)
    
    # Heures supplémentaires par mois
    monthly_overtime = []
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1).date()
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1).date() - timedelta(days=1)
        
        total_overtime = PresenceTracking.objects.filter(
            date__gte=month_start,
            date__lte=month_end
        ).aggregate(total=Sum('overtime_hours'))['total'] or 0
        monthly_overtime.append(float(total_overtime))
    
    # ========== ANALYSES DE CONGÉS ==========
    # Congés par type
    leaves_by_type = LeaveRequest.objects.filter(
        status='RH_APPROVED',
        start_date__year=current_year
    ).values('leave_type').annotate(
        total_days=Sum('days'),
        count=Count('id')
    )
    
    # Congés par mois
    monthly_leaves = []
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1).date()
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1).date() - timedelta(days=1)
        
        month_leaves = LeaveRequest.objects.filter(
            Q(status='MANAGER_APPROVED') | Q(status='RH_APPROVED'),
            start_date__lte=month_end,
            end_date__gte=month_start
        ).count()
        monthly_leaves.append(month_leaves)
    
    # ========== ANALYSES DE RECRUTEMENT ==========
    # Candidats par statut
    candidates_by_status = Candidate.objects.values('status').annotate(
        count=Count('id')
    )
    
    # Taux de conversion (candidats → embauchés)
    total_candidates = Candidate.objects.count()
    hired_candidates = Candidate.objects.filter(status='HIRED').count()
    conversion_rate = (hired_candidates / total_candidates * 100) if total_candidates > 0 else 0
    
    # Entretiens par mois
    monthly_interviews = []
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1)
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1)
        else:
            month_end = datetime(current_year, month + 1, 1)
        
        interviews_count = Interview.objects.filter(
            scheduled_date__gte=month_start,
            scheduled_date__lt=month_end
        ).count()
        monthly_interviews.append(interviews_count)
    
    # ========== ANALYSES DE FORMATION ==========
    # Formations par type
    trainings_by_type = Training.objects.values('training_type').annotate(
        count=Count('id')
    )
    
    # Budget formation
    total_training_budget = TrainingPlan.objects.aggregate(
        total=Sum('budget')
    )['total'] or 0
    
    training_costs = Training.objects.aggregate(
        total=Sum('cost')
    )['total'] or 0
    
    # ========== ANALYSES D'ÉVALUATION ==========
    # Notes moyennes par critère
    avg_scores = Evaluation.objects.filter(
        status='APPROVED',
        evaluation_date__year=current_year
    ).aggregate(
        avg_performance=Avg('performance_score'),
        avg_quality=Avg('quality_score'),
        avg_productivity=Avg('productivity_score'),
        avg_teamwork=Avg('teamwork_score'),
        avg_communication=Avg('communication_score'),
        avg_initiative=Avg('initiative_score')
    )
    
    # ========== TENDANCES ==========
    # Tendance d'embauche (6 derniers mois)
    hiring_trend = []
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30 * i)
        month_hires = Employee.objects.filter(
            date_of_hire__year=month_date.year,
            date_of_hire__month=month_date.month
        ).count()
        hiring_trend.append(month_hires)
    
    return Response({
        'presence_analytics': {
            'monthly_presence_avg': monthly_presence_avg,
            'monthly_lates': monthly_lates,
            'monthly_overtime': monthly_overtime
        },
        'leaves_analytics': {
            'leaves_by_type': list(leaves_by_type),
            'monthly_leaves': monthly_leaves
        },
        'recruitment_analytics': {
            'candidates_by_status': list(candidates_by_status),
            'conversion_rate': round(conversion_rate, 1),
            'monthly_interviews': monthly_interviews,
            'total_candidates': total_candidates,
            'hired_candidates': hired_candidates
        },
        'training_analytics': {
            'trainings_by_type': list(trainings_by_type),
            'total_budget': float(total_training_budget),
            'total_costs': float(training_costs),
            'budget_utilization': (training_costs / total_training_budget * 100) if total_training_budget > 0 else 0
        },
        'evaluation_analytics': {
            'average_scores': {
                'performance': float(avg_scores['avg_performance'] or 0),
                'quality': float(avg_scores['avg_quality'] or 0),
                'productivity': float(avg_scores['avg_productivity'] or 0),
                'teamwork': float(avg_scores['avg_teamwork'] or 0),
                'communication': float(avg_scores['avg_communication'] or 0),
                'initiative': float(avg_scores['avg_initiative'] or 0)
            }
        },
        'trends': {
            'hiring_trend': hiring_trend
        }
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_service_stats(request, service_id=None):
    """Statistiques détaillées par service"""
    from django.db.models import Q, Sum, Avg, Count
    from datetime import datetime, timedelta
    
    today = timezone.now().date()
    current_year = today.year
    
    if service_id:
        try:
            service = Service.objects.get(id=service_id)
        except Service.DoesNotExist:
            return Response(
                {'error': 'Service non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        services = [service]
    else:
        services = Service.objects.all()
    
    service_details = []
    
    for service in services:
        employees = Employee.objects.filter(service=service, is_active=True)
        employee_count = employees.count()
        
        if employee_count == 0:
            continue
        
        # Statistiques de présence
        present_today = PresenceTracking.objects.filter(
            date=today,
            employee__service=service,
            status__in=['PRESENT', 'LATE']
        ).count()
        
        # Congés en cours
        current_leaves = LeaveRequest.objects.filter(
            employee__service=service
        ).filter(
            Q(status='MANAGER_APPROVED') | Q(status='RH_APPROVED')
        ).filter(
            start_date__lte=today,
            end_date__gte=today
        ).count()
        
        # Masse salariale
        service_payroll = Payslip.objects.filter(
            employee__service=service,
            month=today.month,
            year=today.year
        ).aggregate(total=Sum('net_salary'))['total'] or 0
        
        # Contrats expirants
        expiring_contracts = Contract.objects.filter(
            employee__service=service,
            status='SIGNED',
            end_date__lte=today + timedelta(days=90),
            end_date__gte=today
        ).count()
        
        # Formations en cours
        active_trainings = Training.objects.filter(
            employee__service=service,
            status='IN_PROGRESS'
        ).count()
        
        service_details.append({
            'service_id': service.id,
            'service_name': service.name,
            'employee_count': employee_count,
            'present_today': present_today,
            'presence_rate': round((present_today / employee_count * 100), 1) if employee_count > 0 else 0,
            'current_leaves': current_leaves,
            'monthly_payroll': float(service_payroll),
            'avg_salary': float(employees.aggregate(avg=Avg('salary'))['avg'] or 0),
            'expiring_contracts': expiring_contracts,
            'active_trainings': active_trainings
        })
    
    return Response({
        'services': service_details,
        'total_services': len(service_details)
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_alerts(request):
    """Récupérer toutes les alertes du tableau de bord"""
    from django.db.models import Q
    from datetime import timedelta
    
    today = timezone.now().date()
    now = timezone.now()
    
    alerts = {
        'contracts': [],
        'leaves': [],
        'interviews': [],
        'evaluations': [],
        'trainings': []
    }
    
    # Alertes contrats expirants
    expiring_contracts = Contract.objects.filter(
        status='SIGNED',
        end_date__lte=today + timedelta(days=30),
        end_date__gte=today
    ).select_related('employee', 'employee__service').order_by('end_date')
    
    for contract in expiring_contracts:
        alerts['contracts'].append({
            'type': 'contract_expiring',
            'level': contract.alert_level,
            'message': f"Contrat de {contract.employee.get_full_name()} expire dans {contract.days_until_expiry} jour(s)",
            'employee': contract.employee.get_full_name(),
            'end_date': contract.end_date.strftime('%Y-%m-%d'),
            'days_left': contract.days_until_expiry
        })
    
    # Alertes congés en attente depuis plus de 3 jours
    old_pending_leaves = LeaveRequest.objects.filter(
        status='PENDING',
        created_at__lte=now - timedelta(days=3)
    ).select_related('employee')
    
    for leave in old_pending_leaves:
        alerts['leaves'].append({
            'type': 'pending_leave_old',
            'level': 'warning',
            'message': f"Demande de congé de {leave.employee.get_full_name()} en attente depuis plus de 3 jours",
            'employee': leave.employee.get_full_name(),
            'days_pending': (now.date() - leave.created_at.date()).days
        })
    
    # Alertes entretiens aujourd'hui
    today_interviews = Interview.objects.filter(
        scheduled_date__date=today,
        status__in=['SCHEDULED', 'RESCHEDULED']
    ).select_related('candidate', 'interviewer')
    
    for interview in today_interviews:
        alerts['interviews'].append({
            'type': 'interview_today',
            'level': 'info',
            'message': f"Entretien avec {interview.candidate.get_full_name()} aujourd'hui",
            'candidate': interview.candidate.get_full_name(),
            'time': interview.scheduled_date.strftime('%H:%M')
        })
    
    # Alertes évaluations en retard
    from datetime import datetime
    current_year = today.year
    overdue_evaluations = Evaluation.objects.filter(
        evaluation_type='ANNUAL',
        evaluation_date__year=current_year,
        status__in=['DRAFT', 'IN_PROGRESS']
    ).select_related('employee')
    
    for eval in overdue_evaluations:
        days_overdue = (today - eval.evaluation_date).days
        if days_overdue > 0:
            alerts['evaluations'].append({
                'type': 'evaluation_overdue',
                'level': 'warning',
                'message': f"Évaluation annuelle de {eval.employee.get_full_name()} en retard de {days_overdue} jour(s)",
                'employee': eval.employee.get_full_name(),
                'days_overdue': days_overdue
            })
    
    # Compter les alertes par niveau
    critical_count = sum(1 for alert_list in alerts.values() for alert in alert_list if alert.get('level') == 'critical')
    warning_count = sum(1 for alert_list in alerts.values() for alert in alert_list if alert.get('level') == 'warning')
    info_count = sum(1 for alert_list in alerts.values() for alert in alert_list if alert.get('level') == 'info')
    
    return Response({
        'alerts': alerts,
        'summary': {
            'total': sum(len(alert_list) for alert_list in alerts.values()),
            'critical': critical_count,
            'warning': warning_count,
            'info': info_count
        }
    })


class ServiceViewSet(viewsets.ModelViewSet):
    queryset = Service.objects.select_related('manager').all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Optimiser les requêtes avec select_related et annotate pour compter les employés"""
        from django.db.models import Count, Q
        return Service.objects.select_related('manager').annotate(
            employee_count=Count('employees', filter=Q(employees__is_active=True))
        )
    
    def create(self, request, *args, **kwargs):
        """Surcharger create pour mieux gérer les erreurs"""
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            logger.info(f"Données reçues pour création de service: {request.data}")
            return super().create(request, *args, **kwargs)
        except Exception as e:
            logger.error(f"Erreur lors de la création du service: {e}")
            logger.error(f"Données reçues: {request.data}")
            raise


class EmployeeHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet en lecture seule pour l'historique des employés"""
    queryset = EmployeeHistory.objects.all()
    serializer_class = EmployeeHistorySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        employee_id = self.request.query_params.get('employee', None)
        change_type = self.request.query_params.get('change_type', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if change_type:
            queryset = queryset.filter(change_type=change_type)
        
        return queryset.select_related('employee', 'changed_by')


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        service_id = self.request.query_params.get('service', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if service_id:
            queryset = queryset.filter(service_id=service_id)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.select_related('user', 'service')
    
    def perform_create(self, serializer):
        # S'assurer que l'employee_id n'est pas fourni ou sera régénéré s'il ne commence pas par DITECH
        if 'employee_id' in serializer.validated_data:
            employee_id = serializer.validated_data.get('employee_id', '')
            if employee_id and not employee_id.startswith('DITECH'):
                # Supprimer l'ID invalide pour qu'il soit régénéré automatiquement
                del serializer.validated_data['employee_id']
        
        # Créer un User d'abord si l'email est fourni
        email = serializer.validated_data.get('email')
        if email:
            # Générer un username unique à partir de l'email
            base_username = email.split('@')[0]
            username = base_username
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1
            
            user = User.objects.create_user(
                username=username,
                email=email,
                first_name=serializer.validated_data.get('first_name', ''),
                last_name=serializer.validated_data.get('last_name', ''),
                role='EMPLOYE'
            )
            # Créer l'Employee avec ce User (l'ID sera généré automatiquement par la méthode save() si pas fourni)
            employee = serializer.save(user=user)
            
            # Enregistrer la création dans l'historique
            EmployeeHistory.objects.create(
                employee=employee,
                change_type='INFO',
                field_name='creation',
                old_value='',
                new_value='Employé créé',
                description=f'Création de l\'employé {employee.get_full_name()}',
                changed_by=self.request.user
            )
        else:
            # Si pas d'email, laisser le serializer gérer (ne devrait pas arriver normalement)
            serializer.save()
    
    def perform_update(self, serializer):
        # Enregistrer l'utilisateur qui fait la modification pour l'historique
        instance = serializer.instance
        instance._changed_by = self.request.user
        serializer.save()
    
    def perform_destroy(self, instance):
        # Suppression définitive de l'employé
        # Comme Employee.user a on_delete=models.CASCADE, si on supprime le User,
        # l'Employee sera automatiquement supprimé, ainsi que tous les objets liés à l'Employee avec CASCADE
        
        # Récupérer la référence à l'utilisateur avant suppression
        user_to_delete = instance.user if hasattr(instance, 'user') and instance.user else None
        
        # Supprimer directement l'Employee (cela supprimera automatiquement tous les objets liés avec CASCADE)
        # comme EmployeeHistory, Contract, Document, Payslip, etc.
        instance.delete()
        
        # Supprimer l'utilisateur associé pour éviter qu'il reste orphelin dans la base de données
        # Note: Supprimer l'Employee ne supprime pas automatiquement le User (pas de CASCADE inversé)
        if user_to_delete:
            try:
                user_to_delete.delete()
            except Exception as e:
                # Si le User a déjà été supprimé ou s'il y a une erreur, on continue
                # Cela ne devrait normalement pas arriver, mais on gère le cas pour éviter une erreur
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Erreur lors de la suppression du User associé à l'employé: {e}")
    
    @action(detail=True, methods=['get'])
    def dossier(self, request, pk=None):
        """Récupérer le dossier complet d'un employé (informations, contrat, documents, historique)"""
        employee = self.get_object()
        
        # Informations personnelles
        employee_data = EmployeeSerializer(employee).data
        
        # Contrat actif
        active_contract = Contract.objects.filter(
            employee=employee,
            status='SIGNED'
        ).order_by('-start_date').first()
        contract_data = ContractSerializer(active_contract).data if active_contract else None
        
        # Tous les contrats
        all_contracts = Contract.objects.filter(employee=employee).order_by('-start_date')
        contracts_data = ContractSerializer(all_contracts, many=True).data
        
        # Documents
        documents = Document.objects.filter(employee=employee).order_by('-created_at')
        documents_data = DocumentSerializer(documents, many=True, context={'request': request}).data
        
        # Historique des changements
        history = EmployeeHistory.objects.filter(employee=employee).order_by('-changed_at')
        history_data = EmployeeHistorySerializer(history, many=True).data
        
        # Évaluations
        evaluations = Evaluation.objects.filter(employee=employee).order_by('-evaluation_date')
        evaluations_data = EvaluationSerializer(evaluations, many=True).data
        
        # Congés
        leave_requests = LeaveRequest.objects.filter(employee=employee).order_by('-created_at')
        leave_requests_data = LeaveRequestSerializer(leave_requests, many=True).data
        
        # Pointages récents (30 derniers jours)
        from datetime import timedelta
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        attendances = PresenceTracking.objects.filter(
            employee=employee,
            date__gte=thirty_days_ago
        ).order_by('-date')
        attendances_data = PresenceTrackingSerializer(attendances, many=True).data
        
        return Response({
            'employee': employee_data,
            'active_contract': contract_data,
            'all_contracts': contracts_data,
            'documents': documents_data,
            'history': history_data,
            'evaluations': evaluations_data,
            'leave_requests': leave_requests_data,
            'recent_attendances': attendances_data,
        })
    
    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """Récupérer uniquement l'historique des changements d'un employé"""
        employee = self.get_object()
        history = EmployeeHistory.objects.filter(employee=employee).order_by('-changed_at')
        
        # Filtrer par type de changement si fourni
        change_type = request.query_params.get('change_type', None)
        if change_type:
            history = history.filter(change_type=change_type)
        
        serializer = EmployeeHistorySerializer(history, many=True)
        return Response(serializer.data)
    


class JobOfferViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les offres d'emploi"""
    queryset = JobOffer.objects.all()
    serializer_class = JobOfferSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('department', 'created_by')
        status_filter = self.request.query_params.get('status', None)
        department_id = self.request.query_params.get('department', None)
        published_only = self.request.query_params.get('published_only', None)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if published_only and published_only.lower() == 'true':
            queryset = queryset.filter(status='PUBLISHED')
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        """Publier une offre d'emploi"""
        job_offer = self.get_object()
        if job_offer.status != 'DRAFT':
            return Response(
                {'error': 'Seules les offres en brouillon peuvent être publiées'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        job_offer.status = 'PUBLISHED'
        job_offer.published_date = timezone.now()
        job_offer.save()
        
        return Response({
            'message': 'Offre publiée avec succès',
            'job_offer': JobOfferSerializer(job_offer).data
        })
    
    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        """Fermer une offre d'emploi"""
        job_offer = self.get_object()
        job_offer.status = 'CLOSED'
        job_offer.save()
        
        return Response({
            'message': 'Offre fermée avec succès',
            'job_offer': JobOfferSerializer(job_offer).data
        })
    
    @action(detail=False, methods=['get'])
    def open_positions(self, request):
        """Récupérer toutes les offres ouvertes"""
        from django.utils import timezone
        today = timezone.now().date()
        
        open_offers = JobOffer.objects.filter(
            status='PUBLISHED'
        ).select_related('department')
        
        # Filtrer celles qui ne sont pas expirées
        open_offers = [offer for offer in open_offers if offer.is_open]
        
        serializer = self.get_serializer(open_offers, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def applications(self, request, pk=None):
        """Récupérer toutes les candidatures pour une offre"""
        job_offer = self.get_object()
        candidates = Candidate.objects.filter(job_offer=job_offer).order_by('-application_date')
        
        serializer = CandidateSerializer(candidates, many=True)
        return Response({
            'job_offer': JobOfferSerializer(job_offer).data,
            'applications': serializer.data,
            'total': candidates.count()
        })


class CandidateViewSet(viewsets.ModelViewSet):
    queryset = Candidate.objects.all()
    serializer_class = CandidateSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('job_offer', 'job_offer__department', 'created_by')
        status_filter = self.request.query_params.get('status', None)
        job_offer_id = self.request.query_params.get('job_offer', None)
        position = self.request.query_params.get('position', None)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if job_offer_id:
            queryset = queryset.filter(job_offer_id=job_offer_id)
        if position:
            queryset = queryset.filter(position__icontains=position)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Mettre à jour le statut d'un candidat"""
        candidate = self.get_object()
        new_status = request.data.get('status')
        
        if new_status not in dict(Candidate.STATUS_CHOICES):
            return Response(
                {'error': 'Statut invalide'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        candidate.status = new_status
        candidate.save()
        
        return Response({
            'message': 'Statut mis à jour',
            'candidate': CandidateSerializer(candidate).data
        })
    
    @action(detail=True, methods=['post'])
    def hire(self, request, pk=None):
        """Embaucher un candidat (créer un employé)"""
        candidate = self.get_object()
        
        if candidate.status == 'HIRED':
            return Response(
                {'error': 'Ce candidat a déjà été embauché'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create employee from candidate
        service_id = request.data.get('service_id')
        salary = request.data.get('salary', 0)
        date_of_hire = request.data.get('date_of_hire')
        
        if not date_of_hire:
            from datetime import date
            date_of_hire = date.today()
        else:
            from datetime import datetime
            date_of_hire = datetime.strptime(date_of_hire, '%Y-%m-%d').date()
        
        try:
            service = Service.objects.get(id=service_id) if service_id else None
        except Service.DoesNotExist:
            return Response({'error': 'Service not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Create user
        base_username = candidate.email.split('@')[0]
        username = base_username
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{base_username}{counter}"
            counter += 1
        
        user = User.objects.create_user(
            username=username,
            email=candidate.email,
            first_name=candidate.first_name,
            last_name=candidate.last_name,
            role='EMPLOYE'
        )
        
        # Create employee
        last_employee = Employee.objects.filter(employee_id__startswith='DITECH').order_by('employee_id').last()
        if last_employee and last_employee.employee_id.startswith('DITECH'):
            try:
                last_id_num = int(last_employee.employee_id[6:])
                new_id_num = last_id_num + 1
            except (ValueError, IndexError):
                new_id_num = Employee.objects.count() + 1
        else:
            new_id_num = Employee.objects.count() + 1
        
        employee = Employee.objects.create(
            user=user,
            employee_id=f"DITECH{new_id_num:04d}",
            first_name=candidate.first_name,
            last_name=candidate.last_name,
            email=candidate.email,
            phone=candidate.phone,
            date_of_hire=date_of_hire,
            position=candidate.position,
            service=service,
            salary=salary
        )
        
        candidate.status = 'HIRED'
        candidate.save()
        
        from apprh.serializers import EmployeeSerializer
        return Response({
            'message': 'Candidat embauché avec succès',
            'candidate': CandidateSerializer(candidate).data,
            'employee': EmployeeSerializer(employee).data
        })
    
    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        """Récupérer la timeline complète d'un candidat"""
        candidate = self.get_object()
        
        # Récupérer tous les entretiens
        interviews = Interview.objects.filter(candidate=candidate).order_by('scheduled_date')
        
        timeline = []
        
        # Ajouter la candidature
        timeline.append({
            'date': candidate.application_date,
            'type': 'application',
            'title': 'Candidature reçue',
            'description': f'Candidature pour le poste de {candidate.position}',
            'status': candidate.status
        })
        
        # Ajouter les entretiens
        for interview in interviews:
            timeline.append({
                'date': interview.scheduled_date,
                'type': 'interview',
                'title': f'Entretien {interview.get_interview_type_display()}',
                'description': interview.notes or '',
                'status': interview.status,
                'rating': interview.rating,
                'interviewer': interview.interviewer.get_full_name() if interview.interviewer else None
            })
        
        # Trier par date
        timeline.sort(key=lambda x: x['date'])
        
        return Response({
            'candidate': CandidateSerializer(candidate).data,
            'timeline': timeline
        })


class InterviewViewSet(viewsets.ModelViewSet):
    queryset = Interview.objects.all()
    serializer_class = InterviewSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('candidate', 'interviewer')
        candidate_id = self.request.query_params.get('candidate', None)
        interviewer_id = self.request.query_params.get('interviewer', None)
        status_filter = self.request.query_params.get('status', None)
        
        if candidate_id:
            queryset = queryset.filter(candidate_id=candidate_id)
        if interviewer_id:
            queryset = queryset.filter(interviewer_id=interviewer_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.order_by('-scheduled_date')
    
    def perform_create(self, serializer):
        interview = serializer.save()
        # Mettre à jour le statut du candidat
        if interview.candidate.status == 'NEW':
            interview.candidate.status = 'INTERVIEW'
            interview.candidate.save()
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Marquer un entretien comme terminé"""
        interview = self.get_object()
        
        if interview.status == 'COMPLETED':
            return Response(
                {'error': 'Cet entretien est déjà terminé'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Récupérer les données de complétion
        rating = request.data.get('rating', interview.rating)
        feedback = request.data.get('feedback', '')
        strengths = request.data.get('strengths', '')
        weaknesses = request.data.get('weaknesses', '')
        recommendation = request.data.get('recommendation', '')
        duration_minutes = request.data.get('duration_minutes', None)
        
        interview.status = 'COMPLETED'
        interview.actual_date = timezone.now()
        interview.rating = rating
        interview.feedback = feedback
        interview.strengths = strengths
        interview.weaknesses = weaknesses
        if recommendation:
            interview.recommendation = recommendation
        if duration_minutes:
            interview.duration_minutes = duration_minutes
        interview.save()
        
        # Mettre à jour la note globale du candidat (moyenne des entretiens)
        completed_interviews = Interview.objects.filter(
            candidate=interview.candidate,
            status='COMPLETED'
        )
        if completed_interviews.exists():
            avg_rating = completed_interviews.aggregate(
                avg=models.Avg('rating')
            )['avg'] or 0
            interview.candidate.rating = int(round(avg_rating))
            interview.candidate.save()
        
        return Response({
            'message': 'Entretien marqué comme terminé',
            'interview': InterviewSerializer(interview).data
        })
    
    @action(detail=True, methods=['post'])
    def reschedule(self, request, pk=None):
        """Reprogrammer un entretien"""
        interview = self.get_object()
        
        new_date = request.data.get('new_date')
        if not new_date:
            return Response(
                {'error': 'new_date requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from datetime import datetime
        new_date = datetime.fromisoformat(new_date.replace('Z', '+00:00'))
        
        interview.scheduled_date = new_date
        interview.status = 'RESCHEDULED'
        interview.save()
        
        return Response({
            'message': 'Entretien reprogrammé',
            'interview': InterviewSerializer(interview).data
        })
    
    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """Récupérer les entretiens à venir"""
        from django.utils import timezone
        upcoming_interviews = Interview.objects.filter(
            scheduled_date__gte=timezone.now(),
            status__in=['SCHEDULED', 'RESCHEDULED']
        ).select_related('candidate', 'interviewer').order_by('scheduled_date')
        
        serializer = self.get_serializer(upcoming_interviews, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def today(self, request):
        """Récupérer les entretiens d'aujourd'hui"""
        from django.utils import timezone
        from datetime import datetime, timedelta
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        
        today_interviews = Interview.objects.filter(
            scheduled_date__gte=today_start,
            scheduled_date__lt=today_end
        ).select_related('candidate', 'interviewer').order_by('scheduled_date')
        
        serializer = self.get_serializer(today_interviews, many=True)
        return Response(serializer.data)



class LeaveRequestViewSet(viewsets.ModelViewSet):
    queryset = LeaveRequest.objects.all()
    serializer_class = LeaveRequestSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        employee_id = self.request.query_params.get('employee', None)
        status_filter = self.request.query_params.get('status', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def approve_manager(self, request, pk=None):
        """Approuver par le manager"""
        leave_request = self.get_object()
        if leave_request.status != 'PENDING':
            return Response(
                {'error': 'Cette demande n\'est pas en attente'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        leave_request.status = 'MANAGER_APPROVED'
        leave_request.manager_approval = request.user
        leave_request.manager_approval_date = timezone.now()
        leave_request.save()
        
        return Response({
            'status': 'Approuvé par le manager',
            'leave_request': LeaveRequestSerializer(leave_request).data
        })
    
    @action(detail=True, methods=['post'])
    def reject_manager(self, request, pk=None):
        """Rejeter par le manager"""
        leave_request = self.get_object()
        if leave_request.status != 'PENDING':
            return Response(
                {'error': 'Cette demande n\'est pas en attente'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        rejection_reason = request.data.get('rejection_reason', '')
        leave_request.status = 'REJECTED'
        leave_request.manager_rejection_reason = rejection_reason
        leave_request.save()
        
        return Response({
            'status': 'Rejeté par le manager',
            'leave_request': LeaveRequestSerializer(leave_request).data
        })
    
    @action(detail=True, methods=['post'])
    def approve_rh(self, request, pk=None):
        """Approuver par RH et mettre à jour le solde"""
        leave_request = self.get_object()
        if leave_request.status != 'MANAGER_APPROVED':
            return Response(
                {'error': 'La demande doit être approuvée par le manager d\'abord'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Vérifier le solde disponible
        balance, created = LeaveBalance.objects.get_or_create(employee=leave_request.employee)
        
        if leave_request.leave_type == 'ANNUAL':
            if balance.remaining_annual < leave_request.days:
                return Response(
                    {
                        'error': f'Solde insuffisant. Disponible: {balance.remaining_annual} jours, Demandé: {leave_request.days} jours'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        elif leave_request.leave_type == 'SICK':
            if balance.remaining_sick < leave_request.days:
                return Response(
                    {
                        'error': f'Solde maladie insuffisant. Disponible: {balance.remaining_sick} jours, Demandé: {leave_request.days} jours'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Approuver et mettre à jour le solde
        leave_request.status = 'RH_APPROVED'
        leave_request.rh_approval = request.user
        leave_request.rh_approval_date = timezone.now()
        leave_request.save()
        
        # Recalculer le solde (utilise la méthode recalculate_used_days)
        balance.recalculate_used_days()
        
        return Response({
            'status': 'Approuvé par RH et solde mis à jour',
            'leave_request': LeaveRequestSerializer(leave_request).data,
            'updated_balance': LeaveBalanceSerializer(balance).data
        })
    
    @action(detail=True, methods=['post'])
    def reject_rh(self, request, pk=None):
        """Rejeter par RH"""
        leave_request = self.get_object()
        if leave_request.status != 'MANAGER_APPROVED':
            return Response(
                {'error': 'La demande doit être approuvée par le manager d\'abord'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        rejection_reason = request.data.get('rejection_reason', '')
        leave_request.status = 'REJECTED'
        leave_request.rh_rejection_reason = rejection_reason
        leave_request.save()
        
        return Response({
            'status': 'Rejeté par RH',
            'leave_request': LeaveRequestSerializer(leave_request).data
        })
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Annuler une demande de congé"""
        leave_request = self.get_object()
        if leave_request.status == 'RH_APPROVED':
            # Si déjà approuvé, il faut recalculer le solde
            balance, created = LeaveBalance.objects.get_or_create(employee=leave_request.employee)
            balance.recalculate_used_days()
        
        leave_request.status = 'CANCELLED'
        leave_request.save()
        
        return Response({
            'status': 'Demande annulée',
            'leave_request': LeaveRequestSerializer(leave_request).data
        })
    
    @action(detail=False, methods=['get'])
    def current(self, request):
        """Récupérer les congés en cours"""
        today = timezone.now().date()
        current_leaves = LeaveRequest.objects.filter(
            status='RH_APPROVED',
            start_date__lte=today,
            end_date__gte=today
        ).select_related('employee', 'employee__service')
        
        serializer = self.get_serializer(current_leaves, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """Récupérer les congés à venir"""
        today = timezone.now().date()
        upcoming_leaves = LeaveRequest.objects.filter(
            status='RH_APPROVED',
            start_date__gt=today
        ).select_related('employee', 'employee__service').order_by('start_date')
        
        serializer = self.get_serializer(upcoming_leaves, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending_approval(self, request):
        """Récupérer les demandes en attente d'approbation"""
        pending = LeaveRequest.objects.filter(
            status__in=['PENDING', 'MANAGER_APPROVED']
        ).select_related('employee', 'employee__service', 'manager_approval', 'rh_approval')
        
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)


class LeaveBalanceViewSet(viewsets.ModelViewSet):
    queryset = LeaveBalance.objects.all()
    serializer_class = LeaveBalanceSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        employee_id = self.request.query_params.get('employee', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset.select_related('employee')
    
    @action(detail=True, methods=['post'])
    def recalculate(self, request, pk=None):
        """Recalculer automatiquement les soldes de congés"""
        balance = self.get_object()
        balance.recalculate_used_days()
        
        serializer = self.get_serializer(balance)
        return Response({
            'message': 'Soldes recalculés avec succès',
            'balance': serializer.data
        })
    
    @action(detail=False, methods=['post'])
    def recalculate_all(self, request):
        """Recalculer les soldes de tous les employés"""
        balances = LeaveBalance.objects.all()
        updated_count = 0
        
        for balance in balances:
            balance.recalculate_used_days()
            updated_count += 1
        
        return Response({
            'message': f'Soldes recalculés pour {updated_count} employé(s)',
            'updated_count': updated_count
        })


class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        employee_id = self.request.query_params.get('employee', None)
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        return queryset
    

class ContractViewSet(viewsets.ModelViewSet):
    queryset = Contract.objects.all()
    serializer_class = ContractSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('employee', 'employee__service', 'parent_contract', 'created_by')
        employee_id = self.request.query_params.get('employee', None)
        status_filter = self.request.query_params.get('status', None)
        contract_type = self.request.query_params.get('contract_type', None)
        needs_renewal = self.request.query_params.get('needs_renewal', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if contract_type:
            queryset = queryset.filter(contract_type=contract_type)
        if needs_renewal is not None:
            queryset = queryset.filter(needs_renewal=needs_renewal.lower() == 'true')
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def sign_employee(self, request, pk=None):
        contract = self.get_object()
        contract.signed_by_employee = True
        if contract.signed_by_company:
            contract.status = 'SIGNED'
            contract.signed_date = date.today()
        contract.save()
        return Response({'status': 'Signed by employee'})
    
    @action(detail=True, methods=['post'])
    def sign_company(self, request, pk=None):
        contract = self.get_object()
        contract.signed_by_company = True
        if contract.signed_by_employee:
            contract.status = 'SIGNED'
            contract.signed_date = date.today()
        contract.save()
        return Response({'status': 'Signed by company'})
    
    @action(detail=True, methods=['post'])
    def renew(self, request, pk=None):
        """Créer un nouveau contrat en renouvellement de celui-ci"""
        contract = self.get_object()
        
        if contract.contract_type == 'CDI':
            return Response(
                {'error': 'Les contrats CDI ne peuvent pas être renouvelés'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not contract.end_date:
            return Response(
                {'error': 'Ce contrat n\'a pas de date de fin'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Récupérer les données du renouvellement depuis la requête
        new_start_date = request.data.get('new_start_date')
        new_end_date = request.data.get('new_end_date')
        new_salary = request.data.get('new_salary')
        
        if new_start_date:
            from datetime import datetime
            new_start_date = datetime.strptime(new_start_date, '%Y-%m-%d').date()
        if new_end_date:
            from datetime import datetime
            new_end_date = datetime.strptime(new_end_date, '%Y-%m-%d').date()
        if new_salary:
            new_salary = float(new_salary)
        
        try:
            renewal_contract = contract.create_renewal(
                new_start_date=new_start_date,
                new_end_date=new_end_date,
                new_salary=new_salary,
                created_by=request.user
            )
            
            serializer = self.get_serializer(renewal_contract)
            return Response({
                'message': 'Contrat renouvelé avec succès',
                'contract': serializer.data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du renouvellement: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def toggle_auto_renewal(self, request, pk=None):
        """Activer ou désactiver le renouvellement automatique"""
        contract = self.get_object()
        contract.auto_renewal = not contract.auto_renewal
        contract.save()
        
        return Response({
            'message': f'Renouvellement automatique {"activé" if contract.auto_renewal else "désactivé"}',
            'auto_renewal': contract.auto_renewal
        })
    
    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get contracts expiring soon with alert levels"""
        today = date.today()
        days_ahead = int(request.query_params.get('days', 30))
        future_date = today + timedelta(days=days_ahead)
        
        contracts = Contract.objects.filter(
            end_date__gte=today,
            end_date__lte=future_date,
            status='SIGNED'
        ).select_related('employee', 'employee__service')
        
        contracts_data = []
        for contract in contracts:
            contract_dict = self.get_serializer(contract).data
            contract_dict['days_until_expiry'] = contract.days_until_expiry
            contract_dict['alert_level'] = contract.alert_level
            contract_dict['is_expiring_soon'] = contract.is_expiring_soon
            contract_dict['is_expired'] = contract.is_expired
            contracts_data.append(contract_dict)
        
        # Trier par date d'expiration (les plus urgents en premier)
        contracts_data.sort(key=lambda x: x['days_until_expiry'] if x['days_until_expiry'] is not None else float('inf'))
        
        return Response(contracts_data)
    
    @action(detail=False, methods=['get'])
    def expired(self, request):
        """Get expired contracts"""
        today = date.today()
        
        contracts = Contract.objects.filter(
            end_date__lt=today,
            status__in=['SIGNED', 'EXPIRED']
        ).select_related('employee', 'employee__service')
        
        contracts_data = []
        for contract in contracts:
            contract_dict = self.get_serializer(contract).data
            contract_dict['days_until_expiry'] = contract.days_until_expiry
            contract_dict['alert_level'] = 'critical'
            contract_dict['is_expired'] = True
            contracts_data.append(contract_dict)
        
        serializer = self.get_serializer(contracts, many=True)
        return Response(contracts_data)
    
    @action(detail=False, methods=['get'])
    def needs_renewal(self, request):
        """Get contracts that need renewal"""
        contracts = Contract.objects.filter(
            needs_renewal=True,
            status='SIGNED'
        ).select_related('employee', 'employee__service', 'parent_contract')
        
        contracts_data = []
        for contract in contracts:
            contract_dict = self.get_serializer(contract).data
            contract_dict['days_until_expiry'] = contract.days_until_expiry
            contract_dict['alert_level'] = contract.alert_level
            contract_dict['is_expiring_soon'] = contract.is_expiring_soon
            contract_dict['is_expired'] = contract.is_expired
            contracts_data.append(contract_dict)
        
        # Trier par urgence (expirés d'abord, puis par date)
        contracts_data.sort(key=lambda x: (
            x['is_expired'] is False,  # Expirés en premier
            x['days_until_expiry'] if x['days_until_expiry'] is not None else float('inf')
        ))
        
        return Response(contracts_data)
    
    @action(detail=False, methods=['get'])
    def alerts(self, request):
        """Get all contract alerts grouped by level"""
        today = date.today()
        days_ahead = int(request.query_params.get('days', 90))
        future_date = today + timedelta(days=days_ahead)
        
        # Contrats expirant bientôt ou expirés
        contracts = Contract.objects.filter(
            end_date__lte=future_date,
            status='SIGNED'
        ).select_related('employee', 'employee__service')
        
        alerts = {
            'critical': [],
            'warning': [],
            'info': [],
            'expired': []
        }
        
        for contract in contracts:
            if contract.is_expired:
                alerts['expired'].append({
                    'contract': self.get_serializer(contract).data,
                    'days_until_expiry': contract.days_until_expiry,
                    'message': f'Contrat expiré depuis {abs(contract.days_until_expiry)} jour(s)'
                })
            else:
                alert_level = contract.alert_level
                if alert_level:
                    alerts[alert_level].append({
                        'contract': self.get_serializer(contract).data,
                        'days_until_expiry': contract.days_until_expiry,
                        'message': f'Contrat expire dans {contract.days_until_expiry} jour(s)'
                    })
        
        # Compter le total
        total_alerts = sum(len(alerts[key]) for key in alerts)
        
        return Response({
            'alerts': alerts,
            'total': total_alerts,
            'critical_count': len(alerts['critical']) + len(alerts['expired']),
            'warning_count': len(alerts['warning']),
            'info_count': len(alerts['info'])
        })


class PayslipViewSet(viewsets.ModelViewSet):
    queryset = Payslip.objects.all()
    serializer_class = PayslipSerializer
    
    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        # Si le statut est SENT lors de la création, envoyer l'email
        if instance.status == 'SENT':
            self._send_payslip_email(instance)
    
    def perform_update(self, serializer):
        instance = serializer.instance
        old_status = instance.status
        updated_instance = serializer.save()
        # Si le statut est changé à SENT, envoyer l'email
        if updated_instance.status == 'SENT' and old_status != 'SENT':
            self._send_payslip_email(updated_instance)
    
    def _send_payslip_email(self, payslip):
        """Méthode privée pour envoyer l'email de fiche de paie automatiquement"""
        # Utiliser la même logique que la méthode send_email
        # Mais sans retourner de Response (car appelée depuis perform_create/perform_update)
        try:
            employee = payslip.employee
            employee_email = employee.email
            
            if not employee_email:
                print(f"Warning: Employee {employee.get_full_name()} has no email address")
                return
            
            # Préparer les données pour l'email (même logique que send_email)
            month_names = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 
                          'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
            month_name = month_names[payslip.month] if 1 <= payslip.month <= 12 else str(payslip.month)
            
            email_subject = f"Fiche de Paie - {month_name} {payslip.year}"
            
            employee_first_name = str(employee.first_name) if employee.first_name else ''
            employee_last_name = str(employee.last_name) if employee.last_name else ''
            employee_id = str(employee.employee_id) if employee.employee_id else ''
            base_salary = float(payslip.base_salary) if payslip.base_salary else 0.0
            bonuses = float(payslip.bonuses) if payslip.bonuses else 0.0
            deductions = float(payslip.deductions) if payslip.deductions else 0.0
            net_salary = float(payslip.net_salary) if payslip.net_salary else 0.0
            payslip_year = int(payslip.year) if payslip.year else timezone.now().year
            current_year = timezone.now().year
            
            base_salary_formatted = f"{base_salary:,.0f}"
            bonuses_formatted = f"{bonuses:,.0f}"
            deductions_formatted = f"{deductions:,.0f}"
            net_salary_formatted = f"{net_salary:,.0f}"
            
            email_body = """
            <html>
            <head>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        line-height: 1.6;
                        color: #333;
                    }}
                    .container {{
                        max-width: 600px;
                        margin: 0 auto;
                        padding: 20px;
                        background-color: #f9f9f9;
                    }}
                    .header {{
                        background-color: #1e3a8a;
                        color: white;
                        padding: 20px;
                        text-align: center;
                        border-radius: 5px 5px 0 0;
                    }}
                    .content {{
                        background-color: white;
                        padding: 30px;
                        border-radius: 0 0 5px 5px;
                    }}
                    .info-table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 20px 0;
                    }}
                    .info-table td {{
                        padding: 10px;
                        border-bottom: 1px solid #e5e7eb;
                    }}
                    .info-table td:first-child {{
                        font-weight: bold;
                        width: 40%;
                        color: #1e3a8a;
                    }}
                    .salary-table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 20px 0;
                    }}
                    .salary-table th {{
                        background-color: #1e3a8a;
                        color: white;
                        padding: 12px;
                        text-align: left;
                    }}
                    .salary-table td {{
                        padding: 10px;
                        border-bottom: 1px solid #e5e7eb;
                    }}
                    .salary-table tr:last-child {{
                        background-color: #fbbf24;
                        font-weight: bold;
                        font-size: 16px;
                    }}
                    .footer {{
                        margin-top: 30px;
                        padding-top: 20px;
                        border-top: 2px solid #e5e7eb;
                        color: #666;
                        font-size: 12px;
                        text-align: center;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>FICHE DE PAIE</h1>
                        <p>DiTech - Digital Technology Ivoirienne</p>
                    </div>
                    <div class="content">
                        <h2>Bonjour {employee_first_name} {employee_last_name},</h2>
                        <p>Veuillez trouver ci-joint votre fiche de paie pour la période de <strong>{month_name} {payslip_year}</strong>.</p>
                        
                        <table class="info-table">
                            <tr>
                                <td>Employé:</td>
                                <td>{employee_first_name} {employee_last_name}</td>
                            </tr>
                            <tr>
                                <td>ID Employé:</td>
                                <td>{employee_id}</td>
                            </tr>
                            <tr>
                                <td>Période:</td>
                                <td>{month_name} {payslip_year}</td>
                            </tr>
                        </table>
                        
                        <h3>Détails de la paie:</h3>
                        <table class="salary-table">
                            <thead>
                                <tr>
                                    <th>Description</th>
                                    <th style="text-align: right;">Montant (FCFA)</th>
                                </tr>
                            </thead>
                            <tbody>
                                <tr>
                                    <td>Salaire de base</td>
                                    <td style="text-align: right;">{base_salary_formatted}</td>
                                </tr>
                                <tr>
                                    <td>Primes</td>
                                    <td style="text-align: right;">{bonuses_formatted}</td>
                                </tr>
                                <tr>
                                    <td>Déductions</td>
                                    <td style="text-align: right;">-{deductions_formatted}</td>
                                </tr>
                                <tr>
                                    <td>NET À PAYER</td>
                                    <td style="text-align: right;">{net_salary_formatted}</td>
                                </tr>
                            </tbody>
                        </table>
                        
                        <div class="footer">
                            <p>Ceci est un email automatique, merci de ne pas y répondre.</p>
                            <p>Pour toute question, veuillez contacter le service RH.</p>
                            <p>&copy; {current_year} DiTech - Tous droits réservés</p>
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """.format(
                employee_first_name=employee_first_name,
                employee_last_name=employee_last_name,
                employee_id=employee_id,
                month_name=month_name,
                payslip_year=payslip_year,
                base_salary_formatted=base_salary_formatted,
                bonuses_formatted=bonuses_formatted,
                deductions_formatted=deductions_formatted,
                net_salary_formatted=net_salary_formatted,
                current_year=current_year
            )
            
            from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@ditech.com')
            if not from_email or from_email == 'noreply@votredomaine.com':
                from_email = getattr(settings, 'EMAIL_HOST_USER', 'noreply@ditech.com')
            
            email = EmailMessage(
                subject=email_subject,
                body=email_body,
                from_email=from_email,
                to=[employee_email],
            )
            email.content_subtype = "html"
            
            # Attacher le PDF si disponible
            try:
                if payslip.pdf_file and hasattr(payslip.pdf_file, 'path'):
                    pdf_path = payslip.pdf_file.path
                    if os.path.exists(pdf_path):
                        with open(pdf_path, 'rb') as pdf:
                            email.attach(
                                f'Fiche_Paie_{month_name}_{payslip.year}.pdf',
                                pdf.read(),
                                'application/pdf'
                            )
            except Exception as pdf_error:
                print(f"Warning: Could not attach PDF: {str(pdf_error)}")
            
            # Envoyer l'email
            try:
                email.send(fail_silently=False)
                print(f"Email sent successfully to {employee_email}")
                # Mettre à jour la date d'envoi
                payslip.sent_at = timezone.now()
                payslip.save(update_fields=['sent_at'])
            except Exception as send_error:
                import logging
                logger = logging.getLogger(__name__)
                error_message = str(send_error)
                logger.error(f"Failed to send payslip email to {employee_email}: {error_message}")
                print(f"Error sending email: {error_message}")
                # On ne bloque pas la sauvegarde si l'email échoue
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in _send_payslip_email: {str(e)}")
            print(f"Error in _send_payslip_email: {str(e)}")
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('employee', 'employee__service', 'created_by')
        employee_id = self.request.query_params.get('employee', None)
        month = self.request.query_params.get('month', None)
        year = self.request.query_params.get('year', None)
        status_filter = self.request.query_params.get('status', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if month:
            queryset = queryset.filter(month=month)
        if year:
            queryset = queryset.filter(year=year)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def details(self, request, pk=None):
        """Récupérer les détails complets d'une fiche de paie (avec primes et déductions)"""
        payslip = self.get_object()
        
        payslip_data = PayslipSerializer(payslip).data
        bonuses = PayslipBonus.objects.filter(payslip=payslip)
        deductions = PayslipDeduction.objects.filter(payslip=payslip)
        payment_history = PaymentHistory.objects.filter(payslip=payslip)
        
        return Response({
            'payslip': payslip_data,
            'bonuses': PayslipBonusSerializer(bonuses, many=True).data,
            'deductions': PayslipDeductionSerializer(deductions, many=True).data,
            'payment_history': PaymentHistorySerializer(payment_history, many=True).data,
            'summary': {
                'total_bonuses': float(payslip.bonuses),
                'total_deductions': float(payslip.deductions),
                'gross_salary': float(payslip.gross_salary),
                'net_salary': float(payslip.net_salary),
                'total_earnings': float(payslip.total_earnings)
            }
        })
    
    @action(detail=True, methods=['post'])
    def add_bonus(self, request, pk=None):
        """Ajouter une prime à une fiche de paie"""
        payslip = self.get_object()
        
        bonus_data = request.data.copy()
        bonus_data['payslip'] = payslip.id
        
        serializer = PayslipBonusSerializer(data=bonus_data)
        if serializer.is_valid():
            bonus = serializer.save()
            
            # Recalculer le total des primes
            total_bonuses = PayslipBonus.objects.filter(payslip=payslip).aggregate(
                total=models.Sum('amount')
            )['total'] or 0
            payslip.bonuses = total_bonuses
            payslip.save()
            
            return Response({
                'message': 'Prime ajoutée avec succès',
                'bonus': PayslipBonusSerializer(bonus).data,
                'updated_payslip': PayslipSerializer(payslip).data
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def add_deduction(self, request, pk=None):
        """Ajouter une déduction à une fiche de paie"""
        payslip = self.get_object()
        
        deduction_data = request.data.copy()
        deduction_data['payslip'] = payslip.id
        
        serializer = PayslipDeductionSerializer(data=deduction_data)
        if serializer.is_valid():
            deduction = serializer.save()
            
            # Recalculer le total des déductions
            total_deductions = PayslipDeduction.objects.filter(payslip=payslip).aggregate(
                total=models.Sum('amount')
            )['total'] or 0
            payslip.deductions = total_deductions
            payslip.save()
            
            return Response({
                'message': 'Déduction ajoutée avec succès',
                'deduction': PayslipDeductionSerializer(deduction).data,
                'updated_payslip': PayslipSerializer(payslip).data
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def mark_as_paid(self, request, pk=None):
        """Marquer une fiche de paie comme payée"""
        payslip = self.get_object()
        
        payment_date = request.data.get('payment_date')
        payment_method = request.data.get('payment_method', '')
        reference = request.data.get('reference', '')
        notes = request.data.get('notes', '')
        
        if payment_date:
            from datetime import datetime
            payment_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
        else:
            payment_date = date.today()
        
        # Créer un historique de paiement
        payment_history = PaymentHistory.objects.create(
            payslip=payslip,
            payment_date=payment_date,
            amount=payslip.net_salary,
            payment_method=payment_method,
            reference=reference,
            notes=notes,
            created_by=request.user
        )
        
        # Mettre à jour le statut de la fiche de paie
        payslip.status = 'PAID'
        payslip.payment_date = payment_date
        payslip.payment_method = payment_method
        payslip.paid_at = timezone.now()
        payslip.save()
        
        return Response({
            'message': 'Fiche de paie marquée comme payée',
            'payslip': PayslipSerializer(payslip).data,
            'payment_history': PaymentHistorySerializer(payment_history).data
        })
    
    @action(detail=True, methods=['get'])
    def payment_history(self, request, pk=None):
        """Récupérer l'historique des paiements d'une fiche de paie"""
        payslip = self.get_object()
        history = PaymentHistory.objects.filter(payslip=payslip).order_by('-payment_date')
        
        serializer = PaymentHistorySerializer(history, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def employee_history(self, request):
        """Récupérer l'historique des fiches de paie d'un employé"""
        employee_id = request.query_params.get('employee', None)
        
        if not employee_id:
            return Response(
                {'error': 'employee_id requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        payslips = Payslip.objects.filter(employee_id=employee_id).order_by('-year', '-month')
        
        serializer = PayslipSerializer(payslips, many=True)
        return Response({
            'employee_id': employee_id,
            'total_payslips': payslips.count(),
            'payslips': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Statistiques des fiches de paie"""
        from django.db.models import Sum, Avg, Count, Max, Min
        
        employee_id = request.query_params.get('employee', None)
        year = request.query_params.get('year', None)
        
        queryset = Payslip.objects.all()
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if year:
            queryset = queryset.filter(year=year)
        
        stats = queryset.aggregate(
            total_payslips=Count('id'),
            total_net_salary=Sum('net_salary'),
            total_gross_salary=Sum('gross_salary'),
            total_bonuses=Sum('bonuses'),
            total_deductions=Sum('deductions'),
            avg_net_salary=Avg('net_salary'),
            max_net_salary=Max('net_salary'),
            min_net_salary=Min('net_salary')
        )
        
        return Response({
            'period': {'year': year} if year else {'all_years': True},
            'statistics': {
                'total_payslips': stats['total_payslips'] or 0,
                'total_net_salary': float(stats['total_net_salary'] or 0),
                'total_gross_salary': float(stats['total_gross_salary'] or 0),
                'total_bonuses': float(stats['total_bonuses'] or 0),
                'total_deductions': float(stats['total_deductions'] or 0),
                'average_net_salary': float(stats['avg_net_salary'] or 0),
                'max_net_salary': float(stats['max_net_salary'] or 0),
                'min_net_salary': float(stats['min_net_salary'] or 0)
            }
        })
    
    @action(detail=True, methods=['post'])
    def generate_pdf(self, request, pk=None):
        payslip = self.get_object()
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=30,
        )
        title = Paragraph("FICHE DE PAIE", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Employee Info
        employee_info = [
            ['Employé:', f"{payslip.employee.first_name} {payslip.employee.last_name}"],
            ['ID:', payslip.employee.employee_id],
            ['Période:', f"{payslip.month:02d}/{payslip.year}"],
        ]
        
        info_table = Table(employee_info, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Salary Details
        salary_data = [
            ['Description', 'Montant (FCFA)'],
            ['Salaire de base', f"{payslip.base_salary:,.2f}"],
            ['Primes', f"{payslip.bonuses:,.2f}"],
            ['Déductions', f"-{payslip.deductions:,.2f}"],
            ['NET À PAYER', f"{payslip.net_salary:,.2f}"],
        ]
        
        salary_table = Table(salary_data, colWidths=[4*inch, 2*inch])
        salary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fbbf24')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey])
        ]))
        elements.append(salary_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        # Save PDF
        filename = f"payslip_{payslip.employee.employee_id}_{payslip.year}_{payslip.month:02d}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, 'payslips', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        with open(filepath, 'wb') as f:
            f.write(buffer.getvalue())
        
        payslip.pdf_file.name = f"payslips/{filename}"
        payslip.status = 'GENERATED'
        payslip.generated_at = timezone.now()
        payslip.save()
        
        return Response({'message': 'PDF generated successfully', 'file': payslip.pdf_file.url})
    
    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        payslip = self.get_object()
        if not payslip.pdf_file:
            return Response({'error': 'PDF not generated'}, status=status.HTTP_404_NOT_FOUND)
        
        return FileResponse(open(payslip.pdf_file.path, 'rb'), content_type='application/pdf')
    
    @action(detail=True, methods=['post'])
    def send_email(self, request, pk=None):
        payslip = self.get_object()
        if payslip.status != 'GENERATED':
            return Response({'error': 'PDF must be generated first'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Récupérer l'email de l'employé
            employee = payslip.employee
            employee_email = employee.email
            
            if not employee_email:
                return Response({'error': 'Employee email not found'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Préparer les données pour l'email
            month_names = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 
                          'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
            month_name = month_names[payslip.month] if 1 <= payslip.month <= 12 else str(payslip.month)
            
            # Créer le message HTML
            email_subject = f"Fiche de Paie - {month_name} {payslip.year}"
            
            # Échapper les valeurs pour éviter les problèmes dans le HTML
            employee_first_name = str(employee.first_name) if employee.first_name else ''
            employee_last_name = str(employee.last_name) if employee.last_name else ''
            employee_id = str(employee.employee_id) if employee.employee_id else ''
            base_salary = float(payslip.base_salary) if payslip.base_salary else 0.0
            bonuses = float(payslip.bonuses) if payslip.bonuses else 0.0
            deductions = float(payslip.deductions) if payslip.deductions else 0.0
            net_salary = float(payslip.net_salary) if payslip.net_salary else 0.0
            payslip_year = int(payslip.year) if payslip.year else timezone.now().year
            current_year = timezone.now().year
            
            # Formater les montants
            base_salary_formatted = f"{base_salary:,.0f}"
            bonuses_formatted = f"{bonuses:,.0f}"
            deductions_formatted = f"{deductions:,.0f}"
            net_salary_formatted = f"{net_salary:,.0f}"
            
            # Construire le message HTML de manière plus sûre avec format()
            email_body = """
            <html>
            <head>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        line-height: 1.6;
                        color: #333;
                    }}
                    .container {{
                        max-width: 600px;
                        margin: 0 auto;
                        padding: 20px;
                        background-color: #f9f9f9;
                    }}
                    .header {{
                        background-color: #1e3a8a;
                        color: white;
                        padding: 20px;
                        text-align: center;
                        border-radius: 5px 5px 0 0;
                    }}
                    .content {{
                        background-color: white;
                        padding: 30px;
                        border-radius: 0 0 5px 5px;
                    }}
                    .info-table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 20px 0;
                    }}
                    .info-table td {{
                        padding: 10px;
                        border-bottom: 1px solid #e5e7eb;
                    }}
                    .info-table td:first-child {{
                        font-weight: bold;
                        width: 40%;
                        color: #1e3a8a;
                    }}
                    .salary-table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 20px 0;
                    }}
                    .salary-table th {{
                        background-color: #1e3a8a;
                        color: white;
                        padding: 12px;
                        text-align: left;
                    }}
                    .salary-table td {{
                        padding: 10px;
                        border-bottom: 1px solid #e5e7eb;
                    }}
                    .salary-table tr:last-child {{
                        background-color: #fbbf24;
                        font-weight: bold;
                        font-size: 16px;
                    }}
                    .footer {{
                        margin-top: 30px;
                        padding-top: 20px;
                        border-top: 2px solid #e5e7eb;
                        color: #666;
                        font-size: 12px;
                        text-align: center;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>FICHE DE PAIE</h1>
                        <p>DiTech - Digital Technology Ivoirienne</p>
                    </div>
                    <div class="content">
                        <h2>Bonjour {employee_first_name} {employee_last_name},</h2>
                        <p>Veuillez trouver ci-joint votre fiche de paie pour la période de <strong>{month_name} {payslip_year}</strong>.</p>
                        
                        <table class="info-table">
                            <tr>
                                <td>Employé:</td>
                                <td>{employee_first_name} {employee_last_name}</td>
                            </tr>
                            <tr>
                                <td>ID Employé:</td>
                                <td>{employee_id}</td>
                            </tr>
                            <tr>
                                <td>Période:</td>
                                <td>{month_name} {payslip_year}</td>
                            </tr>
                        </table>
                        
                        <h3>Détails de la paie:</h3>
                        <table class="salary-table">
                            <thead>
                                <tr>
                                    <th>Description</th>
                                    <th style="text-align: right;">Montant (FCFA)</th>
                                </tr>
                            </thead>
                            <tbody>
                                <tr>
                                    <td>Salaire de base</td>
                                    <td style="text-align: right;">{base_salary_formatted}</td>
                                </tr>
                                <tr>
                                    <td>Primes</td>
                                    <td style="text-align: right;">{bonuses_formatted}</td>
                                </tr>
                                <tr>
                                    <td>Déductions</td>
                                    <td style="text-align: right;">-{deductions_formatted}</td>
                                </tr>
                                <tr>
                                    <td>NET À PAYER</td>
                                    <td style="text-align: right;">{net_salary_formatted}</td>
                                </tr>
                            </tbody>
                        </table>
                        
                        <div class="footer">
                            <p>Ceci est un email automatique, merci de ne pas y répondre.</p>
                            <p>Pour toute question, veuillez contacter le service RH.</p>
                            <p>&copy; {current_year} DiTech - Tous droits réservés</p>
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """.format(
                employee_first_name=employee_first_name,
                employee_last_name=employee_last_name,
                employee_id=employee_id,
                month_name=month_name,
                payslip_year=payslip_year,
                base_salary_formatted=base_salary_formatted,
                bonuses_formatted=bonuses_formatted,
                deductions_formatted=deductions_formatted,
                net_salary_formatted=net_salary_formatted,
                current_year=current_year
            )
            
            # Créer l'email
            from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@ditech.com')
            if not from_email or from_email == 'noreply@votredomaine.com':
                from_email = getattr(settings, 'EMAIL_HOST_USER', 'noreply@ditech.com')
            
            email = EmailMessage(
                subject=email_subject,
                body=email_body,
                from_email=from_email,
                to=[employee_email],
            )
            email.content_subtype = "html"  # Définir le contenu comme HTML
            
            # Attacher le PDF si disponible
            try:
                if payslip.pdf_file and hasattr(payslip.pdf_file, 'path'):
                    pdf_path = payslip.pdf_file.path
                    if os.path.exists(pdf_path):
                        with open(pdf_path, 'rb') as pdf:
                            email.attach(
                                f'Fiche_Paie_{month_name}_{payslip.year}.pdf',
                                pdf.read(),
                                'application/pdf'
                            )
            except Exception as pdf_error:
                # Si le PDF ne peut pas être attaché, on continue quand même
                print(f"Warning: Could not attach PDF: {str(pdf_error)}")
            
            # Envoyer l'email
            try:
                email.send(fail_silently=False)
                print(f"Email sent successfully to {employee_email}")
            except Exception as send_error:
                # Si l'envoi échoue, on log l'erreur détaillée
                error_message = str(send_error)
                error_str_lower = error_message.lower()
                print(f"Error sending email: {error_message}")
                
                # Vérifier si on est en mode console (développement)
                use_real_email = getattr(settings, 'USE_REAL_EMAIL', False)
                is_console_backend = settings.EMAIL_BACKEND == 'django.core.mail.backends.console.EmailBackend'
                
                if is_console_backend or not use_real_email:
                    # Le backend console affiche l'email dans la console, donc on continue
                    print(f"Email will be displayed in console (development mode)")
                    print(f"To send real emails, set USE_REAL_EMAIL = True in settings.py and configure App Password")
                else:
                    # En mode production avec SMTP, on doit lever l'erreur
                    # Mais on donne un message plus clair
                    if ('application-specific password' in error_str_lower or 
                        'invalidsecondfactor' in error_str_lower or
                        'username and password not accepted' in error_str_lower or
                        'badcredentials' in error_str_lower or
                        '5.7.8' in error_message):
                        # Supprimer le mot de passe de l'erreur avant de la lever
                        raise Exception(
                            "Erreur d'authentification Gmail : Le nom d'utilisateur ou le mot de passe d'application n'est pas valide. "
                            "Vérifiez que :\n"
                            "1. Vous avez créé un 'App Password' (mot de passe d'application) sur https://myaccount.google.com/apppasswords\n"
                            "2. Vous avez activé la validation en deux étapes sur votre compte Gmail\n"
                            "3. Vous avez mis à jour EMAIL_HOST_PASSWORD dans settings.py avec le nouvel App Password (16 caractères sans espaces)\n"
                            "4. USE_REAL_EMAIL est défini à True dans settings.py"
                        )
                    else:
                        raise send_error
            
            # Mettre à jour le statut
            payslip.status = 'SENT'
            payslip.sent_at = timezone.now()
            payslip.save()
            
            return Response({
                'message': f'Fiche de paie envoyée avec succès à {employee_email}',
                'email': employee_email
            })
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            error_message = str(e)
            error_str_lower = error_message.lower()
            print(f"Error sending email: {error_message}")
            print(f"Traceback: {error_trace}")
            
            # Message d'erreur plus clair pour les erreurs d'authentification Gmail
            if ('username and password not accepted' in error_str_lower or
                'badcredentials' in error_str_lower or
                '5.7.8' in error_message or
                'smtpauthenticationerror' in error_str_lower):
                user_friendly_error = (
                    "Erreur d'authentification Gmail : Le mot de passe d'application n'est pas valide.\n\n"
                    "Veuillez :\n"
                    "1. Aller sur https://myaccount.google.com/apppasswords\n"
                    "2. Créer un nouveau mot de passe d'application (App Password)\n"
                    "3. Activer la validation en deux étapes si ce n'est pas déjà fait\n"
                    "4. Mettre à jour EMAIL_HOST_PASSWORD dans settings.py avec le nouveau mot de passe (16 caractères sans espaces)\n"
                    "5. Redémarrer le serveur Django"
                )
            else:
                user_friendly_error = f"Erreur lors de l'envoi de l'email: {error_message}"
            
            return Response({
                'error': user_friendly_error,
                'details': error_message
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_document(request):
    """Endpoint pour uploader un document (scan de fiche de paie, etc.)"""
    try:
        # Vérifier que le fichier est présent
        if 'document' not in request.FILES:
            return Response({
                'error': 'Aucun fichier fourni',
                'detail': 'Le champ "document" est requis'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        document_file = request.FILES['document']
        document_type = request.data.get('document_type', 'OTHER')
        
        # Valider le type de document
        valid_types = [choice[0] for choice in Document.DOCUMENT_TYPE_CHOICES]
        if document_type not in valid_types:
            return Response({
                'error': 'Type de document invalide',
                'detail': f'Le type doit être l\'un des suivants: {", ".join(valid_types)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Valider que c'est une image ou un PDF
        content_type = document_file.content_type or ''
        file_name = document_file.name.lower() if document_file.name else ''
        
        is_image = content_type.startswith('image/')
        is_pdf = (
            content_type == 'application/pdf' or
            file_name.endswith('.pdf')
        )
        
        if not (is_image or is_pdf):
            return Response({
                'error': 'Type de fichier non supporté',
                'detail': 'Seuls les fichiers image (JPG, PNG) et PDF sont acceptés'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Créer le document
        document = Document.objects.create(
            document=document_file,
            document_type=document_type,
            uploaded_by=request.user,
            description=request.data.get('description', '')
        )
        
        # Si c'est un scan de fiche de paie, essayer de lier à un employé si fourni
        if document_type == 'PAYSLIP_SCAN' and request.data.get('employee_id'):
            try:
                employee = Employee.objects.get(id=request.data.get('employee_id'))
                document.employee = employee
                document.save()
            except Employee.DoesNotExist:
                pass  # On continue même si l'employé n'existe pas
        
        return Response({
            'success': True,
            'message': 'Document uploadé avec succès',
            'document': {
                'id': document.id,
                'document_type': document.get_document_type_display(),
                'document_url': document.document.url if document.document else None,
                'created_at': document.created_at,
            }
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({
            'error': 'Erreur lors de l\'upload du document',
            'detail': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def scan_document(request):
    """
    Endpoint pour numériser un document PDF ou une image (OCR et extraction de données)
    POST /ditech/documents/scan/
    
    Body (FormData):
    - file: fichier PDF ou image (JPG, PNG, GIF, BMP, TIFF, WEBP) à numériser - max 10 MB
    - document_type: type de document (optionnel, défaut: 'employee_document')
    
    Formats supportés:
    - PDF: Extraction de texte avec PyPDF2, OCR optionnel si peu de texte
    - Images (JPG, PNG, etc.): OCR direct avec pytesseract
    
    Retourne:
    - Texte extrait (OCR pour les images, PyPDF2 puis OCR pour les PDFs)
    - Données structurées (emails, téléphones, dates, adresses, etc.)
    - Détection automatique du type de document (passeport, carte d'identité, formulaire, etc.)
    - Métadonnées du document
    """
    import logging
    import re
    import tempfile
    import os
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    
    logger = logging.getLogger(__name__)
    
    try:
        # Vérifier qu'un fichier a été uploadé
        if 'file' not in request.FILES:
            return Response(
                {'detail': 'Aucun fichier fourni. Le champ "file" est requis.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        uploaded_file = request.FILES['file']
        document_type = request.POST.get('document_type', 'employee_document')
        
        # Vérifier le type de fichier (PDF ou images)
        file_name = uploaded_file.name.lower()
        content_type = uploaded_file.content_type or ''
        
        is_pdf = file_name.endswith('.pdf') or content_type == 'application/pdf'
        is_image = (
            file_name.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')) or
            content_type.startswith('image/')
        )
        
        if not (is_pdf or is_image):
            return Response(
                {'detail': 'Le fichier doit être un PDF ou une image (JPG, PNG, etc.)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Vérifier la taille (max 10MB)
        if uploaded_file.size > 10 * 1024 * 1024:
            return Response(
                {'detail': 'Le fichier ne doit pas dépasser 10 MB'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Initialiser le résultat
        result = {
            'document_info': {
                'filename': uploaded_file.name,
                'document_type': document_type,
                'file_size': uploaded_file.size,
                'scanned_at': timezone.now().isoformat(),
                'scanned_by': request.user.username if hasattr(request.user, 'username') else None
            },
            'extracted_text': '',
            'structured_data': {},
            'metadata': {}
        }
        
        # Sauvegarder temporairement le fichier
        file_extension = os.path.splitext(uploaded_file.name)[1].lower()
        file_path = default_storage.save(
            f'temp_scan_{timezone.now().strftime("%Y%m%d_%H%M%S")}_{uploaded_file.name}',
            ContentFile(uploaded_file.read())
        )
        full_path = default_storage.path(file_path)
        
        try:
            # Si c'est une image, utiliser directement l'OCR
            if is_image:
                result['document_info']['file_type'] = 'image'
                result['document_info']['image_format'] = file_extension.replace('.', '').upper()
                
                # Pour les images, on utilise directement l'OCR
                try:
                    from PIL import Image
                    import pytesseract
                    
                    # Ouvrir l'image
                    image = Image.open(full_path)
                    
                    # Obtenir les dimensions
                    result['document_info']['image_width'] = image.width
                    result['document_info']['image_height'] = image.height
                    result['document_info']['pages'] = 1
                    
                    # Effectuer l'OCR
                    ocr_text = pytesseract.image_to_string(image, lang='fra+eng')
                    
                    if ocr_text.strip():
                        result['extracted_text'] = ocr_text
                        result['document_info']['ocr_used'] = True
                        result['document_info']['text_extraction_method'] = 'OCR'
                    else:
                        result['extracted_text'] = "Aucun texte détecté dans l'image"
                        result['document_info']['ocr_used'] = True
                        result['document_info']['text_extraction_method'] = 'OCR (aucun texte)'
                        
                except ImportError:
                    logger.warning("pytesseract ou PIL n'est pas installé. OCR non disponible pour les images.")
                    result['extracted_text'] = "OCR non disponible. Veuillez installer pytesseract et PIL (Pillow)."
                    result['metadata']['error'] = 'OCR libraries not installed'
                except Exception as e:
                    logger.error(f"Erreur lors de l'OCR de l'image: {e}")
                    result['extracted_text'] = f"Erreur lors de l'OCR: {str(e)}"
                    result['metadata']['ocr_error'] = str(e)
            
            # Si c'est un PDF, utiliser PyPDF2 puis OCR si nécessaire
            elif is_pdf:
                result['document_info']['file_type'] = 'pdf'
                
                # Essayer d'utiliser PyPDF2 pour extraire le texte
                try:
                    import PyPDF2
                    
                    with open(full_path, 'rb') as pdf_file:
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        result['document_info']['pages'] = len(pdf_reader.pages)
                        
                        # Extraire les métadonnées
                        if pdf_reader.metadata:
                            result['metadata'] = {
                                'title': str(pdf_reader.metadata.get('/Title', '')),
                                'author': str(pdf_reader.metadata.get('/Author', '')),
                                'subject': str(pdf_reader.metadata.get('/Subject', '')),
                                'creator': str(pdf_reader.metadata.get('/Creator', '')),
                                'producer': str(pdf_reader.metadata.get('/Producer', '')),
                                'creation_date': str(pdf_reader.metadata.get('/CreationDate', '')),
                                'modification_date': str(pdf_reader.metadata.get('/ModDate', ''))
                            }
                        
                        # Extraire le texte de toutes les pages
                        extracted_text_parts = []
                        for page_num, page in enumerate(pdf_reader.pages, 1):
                            try:
                                text = page.extract_text()
                                if text:
                                    extracted_text_parts.append(f"=== Page {page_num} ===\n{text}")
                            except Exception as e:
                                logger.warning(f"Erreur lors de l'extraction du texte de la page {page_num}: {e}")
                        
                        result['extracted_text'] = '\n\n'.join(extracted_text_parts)
                        result['document_info']['text_extraction_method'] = 'PyPDF2'
                        
                except ImportError:
                    logger.warning("PyPDF2 n'est pas installé. Installation: pip install PyPDF2")
                    result['metadata']['error'] = 'PyPDF2 n\'est pas installé'
                    result['extracted_text'] = 'PyPDF2 n\'est pas installé. Extraction de texte impossible.'
                except Exception as e:
                    logger.error(f"Erreur lors de la lecture du PDF avec PyPDF2: {e}")
                    result['extracted_text'] = f"Erreur lors de l'extraction du texte: {str(e)}"
            
            # OCR optionnel pour PDFs (si peu de texte extrait et si pytesseract est disponible)
            if is_pdf and (not result.get('extracted_text') or len(result.get('extracted_text', '').strip()) < 50):
                try:
                    from pdf2image import convert_from_path
                    import pytesseract
                    
                    # Convertir le PDF en images
                    images = convert_from_path(full_path, dpi=300)
                    ocr_text_parts = []
                    
                    for i, image in enumerate(images, 1):
                        try:
                            # Effectuer l'OCR sur chaque page
                            ocr_text = pytesseract.image_to_string(image, lang='fra+eng')
                            if ocr_text.strip():
                                ocr_text_parts.append(f"=== Page {i} (OCR) ===\n{ocr_text}")
                        except Exception as e:
                            logger.warning(f"Erreur OCR pour la page {i}: {e}")
                    
                    if ocr_text_parts:
                        ocr_result = '\n\n'.join(ocr_text_parts)
                        if result['extracted_text']:
                            result['extracted_text'] += f"\n\n=== Texte OCR supplémentaire ===\n{ocr_result}"
                        else:
                            result['extracted_text'] = ocr_result
                        result['document_info']['ocr_used'] = True
                        
                except ImportError:
                    logger.info("pytesseract ou pdf2image n'est pas installé. OCR non disponible.")
                except Exception as e:
                    logger.error(f"Erreur lors de l'OCR: {e}")
                    result['metadata']['ocr_error'] = str(e)
            
            # Extraire des données structurées du texte
            if result['extracted_text']:
                text = result['extracted_text']
                text_lower = text.lower()
                structured_data = {
                    'type': document_type,
                    'extracted_fields': {},
                    'document_specific': {}
                }
                
                # Détecter le type de document basé sur le contenu
                is_passport = 'passeport' in text_lower or 'passport' in text_lower or 'pass' in text_lower[:50]
                is_id_card = 'carte' in text_lower and ('identité' in text_lower or 'identite' in text_lower or 'national' in text_lower)
                is_diploma = 'diplôme' in text_lower or 'diplome' in text_lower or 'certificat' in text_lower
                is_address_change = 'changement' in text_lower and ('adresse' in text_lower or 'address' in text_lower)
                is_form = 'formulaire' in text_lower or 'form' in text_lower[:30]
                
                # Patterns pour extraire des informations générales
                # Emails
                email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                emails = re.findall(email_pattern, text)
                if emails:
                    structured_data['extracted_fields']['emails'] = list(set(emails))[:5]
                
                # Téléphones (amélioré)
                phone_pattern = r'(?:\+?\d{1,3}[\s\-]?)?\(?\d{2,3}\)?[\s\-]?\d{2}[\s\-]?\d{2}[\s\-]?\d{2}[\s\-]?\d{2,3}'
                phones = re.findall(phone_pattern, text)
                if phones:
                    structured_data['extracted_fields']['telephones'] = list(set(phones))[:5]
                
                # Dates (formats variés)
                date_patterns = [
                    r'\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}',  # DD/MM/YYYY ou DD-MM-YYYY
                    r'\d{4}[/\-]\d{1,2}[/\-]\d{1,2}',     # YYYY/MM/DD
                    r'\d{1,2}\s+(?:janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s+\d{2,4}',
                    r'(?:janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s+\d{1,2},?\s+\d{2,4}'
                ]
                all_dates = []
                for pattern in date_patterns:
                    dates = re.findall(pattern, text, re.IGNORECASE)
                    all_dates.extend(dates)
                if all_dates:
                    structured_data['extracted_fields']['dates'] = list(set(all_dates))[:15]
                
                # Extraction spécifique pour les passeports
                if is_passport:
                    # Numéro de passeport (format variable: lettres et chiffres, souvent 8-9 caractères)
                    passport_number_pattern = r'(?:passeport|passport|pass|numéro|n°|no)[\s:]*([A-Z]{1,2}\d{6,9}|\d{8,10}|[A-Z0-9]{8,12})'
                    passport_num_match = re.search(passport_number_pattern, text, re.IGNORECASE)
                    if passport_num_match:
                        structured_data['document_specific']['passport_number'] = passport_num_match.group(1).strip()
                    
                    # Numéro MRZ (Machine Readable Zone) - ligne en bas du passeport
                    mrz_pattern = r'[A-Z0-9<]{44}|[A-Z]{2}[A-Z0-9<]{9}\d[A-Z0-9<]{15}'
                    mrz = re.findall(mrz_pattern, text)
                    if mrz:
                        structured_data['document_specific']['mrz'] = mrz[:2]  # Généralement 2 lignes
                    
                    # Nationalité
                    nationality_pattern = r'(?:nationalité|nationality|national)[\s:]*([A-Z]{2,3}|[A-Z][a-z]+)'
                    nationality_match = re.search(nationality_pattern, text, re.IGNORECASE)
                    if nationality_match:
                        structured_data['document_specific']['nationality'] = nationality_match.group(1).strip()
                    
                    # Date de naissance (chercher après "né le", "born", "date of birth")
                    birth_date_pattern = r'(?:né le|born|date of birth|naissance|date de naissance)[\s:]*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'
                    birth_date_match = re.search(birth_date_pattern, text, re.IGNORECASE)
                    if birth_date_match:
                        structured_data['document_specific']['birth_date'] = birth_date_match.group(1).strip()
                    
                    # Date d'expiration
                    expiry_pattern = r'(?:expire|expiration|valid until|valable jusqu|date d[^\s]*expiration)[\s:]*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'
                    expiry_match = re.search(expiry_pattern, text, re.IGNORECASE)
                    if expiry_match:
                        structured_data['document_specific']['expiry_date'] = expiry_match.group(1).strip()
                    
                    # Lieu de naissance
                    birth_place_pattern = r'(?:né à|born in|lieu de naissance|birth place)[\s:]*([A-Z][^,\n]+)'
                    birth_place_match = re.search(birth_place_pattern, text, re.IGNORECASE)
                    if birth_place_match:
                        structured_data['document_specific']['birth_place'] = birth_place_match.group(1).strip()[:100]
                    
                    structured_data['document_type_detected'] = 'passport'
                
                # Extraction spécifique pour les cartes d'identité
                elif is_id_card:
                    # Numéro de carte d'identité
                    id_number_pattern = r'(?:numéro|n°|no|number)[\s:]*([A-Z0-9]{8,12})'
                    id_num_match = re.search(id_number_pattern, text, re.IGNORECASE)
                    if id_num_match:
                        structured_data['document_specific']['id_number'] = id_num_match.group(1).strip()
                    
                    structured_data['document_type_detected'] = 'id_card'
                
                # Extraction spécifique pour les formulaires de changement d'adresse
                elif is_address_change:
                    # Ancienne adresse
                    old_address_pattern = r'(?:ancienne|old|précédente|précedente|ancien)[\s]*(?:adresse|address)[\s:]*([^\n]{10,200})'
                    old_addr_match = re.search(old_address_pattern, text, re.IGNORECASE)
                    if old_addr_match:
                        structured_data['document_specific']['old_address'] = old_addr_match.group(1).strip()[:200]
                    
                    # Nouvelle adresse
                    new_address_pattern = r'(?:nouvelle|new|nouveau)[\s]*(?:adresse|address)[\s:]*([^\n]{10,200})'
                    new_addr_match = re.search(new_address_pattern, text, re.IGNORECASE)
                    if new_addr_match:
                        structured_data['document_specific']['new_address'] = new_addr_match.group(1).strip()[:200]
                    
                    # Date de changement
                    change_date_pattern = r'(?:date|le)[\s]*(?:de|du)[\s]*(?:changement|déménagement|deménagement)[\s:]*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'
                    change_date_match = re.search(change_date_pattern, text, re.IGNORECASE)
                    if change_date_match:
                        structured_data['document_specific']['change_date'] = change_date_match.group(1).strip()
                    
                    # Raison du changement (si mentionnée)
                    reason_pattern = r'(?:raison|motif|reason)[\s:]*([^\n]{10,150})'
                    reason_match = re.search(reason_pattern, text, re.IGNORECASE)
                    if reason_match:
                        structured_data['document_specific']['reason'] = reason_match.group(1).strip()[:150]
                    
                    structured_data['document_type_detected'] = 'address_change_form'
                
                # Détection générale de formulaires
                elif is_form:
                    structured_data['document_type_detected'] = 'form'
                
                # Extraction de noms (pour tous les types de documents)
                # Nom complet (mots en majuscules consécutifs)
                name_pattern = r'\b([A-ZÀÁÂÃÄÅÆÇÈÉÊËÌÍÎÏÐÑÒÓÔÕÖØÙÚÛÜÝÞ][A-ZÀÁÂÃÄÅÆÇÈÉÊËÌÍÎÏÐÑÒÓÔÕÖØÙÚÛÜÝÞ\s]{2,})\b'
                potential_names = re.findall(name_pattern, text)
                if potential_names:
                    # Filtrer les faux positifs (trop courts, mots communs)
                    filtered_names = [n.strip() for n in potential_names if len(n.strip()) >= 3 and 
                                     n.strip().upper() not in ['MR', 'MRS', 'MLLE', 'MONSIEUR', 'MADAME', 'PASSEPORT', 'PASSPORT']]
                    if filtered_names:
                        structured_data['extracted_fields']['potential_names'] = list(set(filtered_names))[:10]
                
                # Adresses (lignes contenant des numéros de rue)
                address_pattern = r'\d+[,\s]+(?:rue|avenue|boulevard|route|impasse|chemin|place)[\s,]+[A-ZÀÁÂÃÄÅÆÇÈÉÊËÌÍÎÏÐÑÒÓÔÕÖØÙÚÛÜÝÞ][^,\n]+'
                addresses = re.findall(address_pattern, text, re.IGNORECASE)
                if addresses:
                    structured_data['extracted_fields']['addresses'] = [a.strip()[:200] for a in addresses[:5]]
                
                # Statistiques
                structured_data['statistics'] = {
                    'total_characters': len(text),
                    'total_words': len(text.split()),
                    'total_lines': len(text.split('\n')),
                    'has_email': bool(structured_data['extracted_fields'].get('emails')),
                    'has_phone': bool(structured_data['extracted_fields'].get('telephones')),
                    'has_dates': bool(structured_data['extracted_fields'].get('dates')),
                    'has_names': bool(structured_data['extracted_fields'].get('potential_names')),
                    'has_addresses': bool(structured_data['extracted_fields'].get('addresses')),
                    'document_category': structured_data.get('document_type_detected', 'generic')
                }
                
                result['structured_data'] = structured_data
            
            # Sauvegarder le document dans la base de données
            saved_document = None
            try:
                # Lire le fichier depuis le chemin temporaire
                with open(full_path, 'rb') as file_to_save:
                    file_content = file_to_save.read()
                
                # Déterminer le type de document à partir du nom ou de la détection
                doc_type_choice = 'OTHER'
                if result.get('structured_data', {}).get('document_type_detected') == 'passport':
                    doc_type_choice = 'ID_CARD'  # Utiliser ID_CARD pour les passeports
                elif result.get('structured_data', {}).get('document_type_detected') == 'id_card':
                    doc_type_choice = 'ID_CARD'
                
                # Créer la description avec les informations extraites
                description_parts = []
                if result.get('structured_data', {}).get('document_specific'):
                    for key, value in result['structured_data']['document_specific'].items():
                        if value:
                            description_parts.append(f"{key}: {value}")
                
                description = f"Document numérisé - {result.get('structured_data', {}).get('document_type_detected', 'générique')}"
                if description_parts:
                    description += f"\nDonnées extraites: {', '.join(description_parts[:3])}"
                
                # Créer le document dans la base de données
                saved_document = Document.objects.create(
                    document=ContentFile(
                        file_content,
                        name=uploaded_file.name
                    ),
                    document_type=doc_type_choice,
                    uploaded_by=request.user,
                    description=description[:500]  # Limiter à 500 caractères
                )
                
                # Si un employee_id est fourni, lier le document à l'employé
                employee_id = request.POST.get('employee_id')
                if employee_id:
                    try:
                        employee = Employee.objects.get(id=employee_id)
                        saved_document.employee = employee
                        saved_document.save()
                        result['document_info']['linked_to_employee'] = {
                            'id': employee.id,
                            'name': employee.get_full_name(),
                            'employee_id': employee.employee_id
                        }
                    except Employee.DoesNotExist:
                        logger.warning(f"Employé avec l'ID {employee_id} non trouvé")
                
                result['document_id'] = saved_document.id
                result['document_url'] = saved_document.document.url if saved_document.document else None
                result['document_info']['saved'] = True
                result['document_info']['saved_at'] = saved_document.created_at.isoformat()
                
            except Exception as e:
                logger.error(f"Erreur lors de la sauvegarde du document: {e}")
                import traceback
                logger.error(traceback.format_exc())
                result['document_info']['save_error'] = str(e)
                result['document_info']['saved'] = False
            
            return Response(result, status=status.HTTP_200_OK)
            
        finally:
            # Nettoyer le fichier temporaire
            try:
                if default_storage.exists(file_path):
                    default_storage.delete(file_path)
            except Exception as e:
                logger.error(f"Erreur lors de la suppression du fichier temporaire: {e}")
                
    except Exception as e:
        logger.error(f"Erreur lors de la numérisation du document: {e}")
        import traceback
        return Response(
            {'detail': f'Erreur lors de la numérisation: {str(e)}', 'traceback': traceback.format_exc()},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class PayslipBonusViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les primes des fiches de paie"""
    queryset = PayslipBonus.objects.all()
    serializer_class = PayslipBonusSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('payslip', 'payslip__employee')
        payslip_id = self.request.query_params.get('payslip', None)
        
        if payslip_id:
            queryset = queryset.filter(payslip_id=payslip_id)
        
        return queryset
    
    def perform_create(self, serializer):
        bonus = serializer.save()
        # Recalculer le total des primes de la fiche de paie
        payslip = bonus.payslip
        total_bonuses = PayslipBonus.objects.filter(payslip=payslip).aggregate(
            total=models.Sum('amount')
        )['total'] or 0
        payslip.bonuses = total_bonuses
        payslip.save()
    
    def perform_destroy(self, instance):
        payslip = instance.payslip
        super().perform_destroy(instance)
        # Recalculer le total des primes après suppression
        total_bonuses = PayslipBonus.objects.filter(payslip=payslip).aggregate(
            total=models.Sum('amount')
        )['total'] or 0
        payslip.bonuses = total_bonuses
        payslip.save()


class PayslipDeductionViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les déductions des fiches de paie"""
    queryset = PayslipDeduction.objects.all()
    serializer_class = PayslipDeductionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('payslip', 'payslip__employee')
        payslip_id = self.request.query_params.get('payslip', None)
        
        if payslip_id:
            queryset = queryset.filter(payslip_id=payslip_id)
        
        return queryset
    
    def perform_create(self, serializer):
        deduction = serializer.save()
        # Recalculer le total des déductions de la fiche de paie
        payslip = deduction.payslip
        total_deductions = PayslipDeduction.objects.filter(payslip=payslip).aggregate(
            total=models.Sum('amount')
        )['total'] or 0
        payslip.deductions = total_deductions
        payslip.save()
    
    def perform_destroy(self, instance):
        payslip = instance.payslip
        super().perform_destroy(instance)
        # Recalculer le total des déductions après suppression
        total_deductions = PayslipDeduction.objects.filter(payslip=payslip).aggregate(
            total=models.Sum('amount')
        )['total'] or 0
        payslip.deductions = total_deductions
        payslip.save()


class PaymentHistoryViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer l'historique des paiements"""
    queryset = PaymentHistory.objects.all()
    serializer_class = PaymentHistorySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('payslip', 'payslip__employee', 'created_by')
        payslip_id = self.request.query_params.get('payslip', None)
        employee_id = self.request.query_params.get('employee', None)
        
        if payslip_id:
            queryset = queryset.filter(payslip_id=payslip_id)
        if employee_id:
            queryset = queryset.filter(payslip__employee_id=employee_id)
        
        return queryset.order_by('-payment_date', '-created_at')
    
    def perform_create(self, serializer):
        payment = serializer.save(created_by=self.request.user)
        # Mettre à jour le statut de la fiche de paie si nécessaire
        payslip = payment.payslip
        if payslip.status != 'PAID':
            payslip.status = 'PAID'
            payslip.payment_date = payment.payment_date
            payslip.payment_method = payment.payment_method
            payslip.paid_at = timezone.now()
            payslip.save()


class DocumentViewSet(viewsets.ModelViewSet):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        document_type = self.request.query_params.get('document_type', None)
        employee_id = self.request.query_params.get('employee_id', None)
        
        if document_type:
            queryset = queryset.filter(document_type=document_type)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        # Trier par date de création décroissante
        queryset = queryset.order_by('-created_at')
        
        return queryset
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)


class PresenceTrackingViewSet(viewsets.ModelViewSet):
    queryset = PresenceTracking.objects.all()
    serializer_class = PresenceTrackingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        employee_id = self.request.query_params.get('employee', None)
        date_filter = self.request.query_params.get('date', None)
        status_filter = self.request.query_params.get('status', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if date_filter:
            queryset = queryset.filter(date=date_filter)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.order_by('-date', '-check_in_time')
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=False, methods=['post'])
    def check_in(self, request):
        """Pointage d'arrivée (check-in)"""
        from datetime import datetime
        
        employee_id = request.data.get('employee_id')
        badge_id = request.data.get('badge_id')
        check_in_method = request.data.get('check_in_method', 'MANUAL')
        
        if not employee_id and not badge_id:
            return Response(
                {'error': 'employee_id ou badge_id requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Trouver l'employé
        try:
            if badge_id:
                employee = Employee.objects.get(badge_id=badge_id)
            else:
                employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employé non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        today = timezone.now().date()
        now = timezone.now()
        
        # Vérifier si un pointage existe déjà pour aujourd'hui
        tracking, created = PresenceTracking.objects.get_or_create(
            employee=employee,
            date=today,
            defaults={
                'check_in_time': now,
                'check_in_method': check_in_method,
                'badge_id': badge_id if badge_id else '',
                'status': 'PRESENT',
                'created_by': request.user
            }
        )
        
        if not created:
            # Mettre à jour le pointage existant
            if not tracking.check_in_time:
                tracking.check_in_time = now
                tracking.check_in_method = check_in_method
                if badge_id:
                    tracking.badge_id = badge_id
                tracking.status = 'PRESENT'
                tracking.save()
            else:
                return Response(
                    {'error': 'Pointage d\'arrivée déjà enregistré pour aujourd\'hui'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        serializer = self.get_serializer(tracking)
        return Response({
            'message': 'Pointage d\'arrivée enregistré',
            'tracking': serializer.data
        })
    
    @action(detail=False, methods=['post'])
    def check_out(self, request):
        """Pointage de départ (check-out)"""
        employee_id = request.data.get('employee_id')
        badge_id = request.data.get('badge_id')
        
        if not employee_id and not badge_id:
            return Response(
                {'error': 'employee_id ou badge_id requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Trouver l'employé
        try:
            if badge_id:
                employee = Employee.objects.get(badge_id=badge_id)
            else:
                employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Employé non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        today = timezone.now().date()
        now = timezone.now()
        
        try:
            tracking = PresenceTracking.objects.get(employee=employee, date=today)
            
            if tracking.check_out_time:
                return Response(
                    {'error': 'Pointage de départ déjà enregistré pour aujourd\'hui'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            tracking.check_out_time = now
            tracking.save()
            
            serializer = self.get_serializer(tracking)
            return Response({
                'message': 'Pointage de départ enregistré',
                'tracking': serializer.data
            })
        except PresenceTracking.DoesNotExist:
            return Response(
                {'error': 'Aucun pointage d\'arrivée trouvé pour aujourd\'hui'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['post'])
    def badge_check_in(self, request):
        """Pointage avec badge (check-in automatique)"""
        badge_id = request.data.get('badge_id')
        
        if not badge_id:
            return Response(
                {'error': 'badge_id requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Trouver l'employé par badge_id
        try:
            employee = Employee.objects.get(badge_id=badge_id)
        except Employee.DoesNotExist:
            return Response(
                {'error': 'Badge non reconnu'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Utiliser la méthode check_in avec badge
        request.data['employee_id'] = employee.id
        request.data['check_in_method'] = 'BADGE'
        return self.check_in(request)
    
    @action(detail=False, methods=['get'])
    def present_today(self, request):
        """Récupérer tous les employés présents aujourd'hui"""
        today = timezone.now().date()
        present_trackings = PresenceTracking.objects.filter(
            date=today,
            status__in=['PRESENT', 'LATE']
        ).select_related('employee', 'employee__service')
        
        present_employees = []
        for tracking in present_trackings:
            present_employees.append({
                'id': tracking.employee.id,
                'first_name': tracking.employee.first_name,
                'last_name': tracking.employee.last_name,
                'employee_id': tracking.employee.employee_id,
                'service_name': tracking.employee.service.name if tracking.employee.service else 'Non assigné',
                'department': tracking.employee.service.name if tracking.employee.service else 'Non assigné',
                'check_in_time': tracking.check_in_time.isoformat() if tracking.check_in_time else None,
                'check_out_time': tracking.check_out_time.isoformat() if tracking.check_out_time else None,
                'status': tracking.status,
                'is_late': tracking.is_late,
                'late_minutes': tracking.late_minutes,
                'worked_hours': float(tracking.worked_hours),
                'overtime_hours': float(tracking.overtime_hours),
                'check_in_method': tracking.check_in_method
            })
        
        return Response(present_employees)
    
    @action(detail=False, methods=['get'])
    def late_employees(self, request):
        """Récupérer les employés en retard"""
        date_filter = request.query_params.get('date', None)
        if date_filter:
            from datetime import datetime
            target_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
        else:
            target_date = timezone.now().date()
        
        late_trackings = PresenceTracking.objects.filter(
            date=target_date,
            is_late=True
        ).select_related('employee', 'employee__service')
        
        late_employees = []
        for tracking in late_trackings:
            late_employees.append({
                'employee': {
                    'id': tracking.employee.id,
                    'name': tracking.employee.get_full_name(),
                    'employee_id': tracking.employee.employee_id,
                    'service': tracking.employee.service.name if tracking.employee.service else 'Non assigné'
                },
                'check_in_time': tracking.check_in_time.isoformat() if tracking.check_in_time else None,
                'late_minutes': tracking.late_minutes,
                'expected_check_in': tracking.expected_check_in.strftime('%H:%M') if tracking.expected_check_in else None
            })
        
        return Response({
            'date': target_date,
            'total_late': len(late_employees),
            'employees': late_employees
        })
    
    @action(detail=False, methods=['get'])
    def overtime_stats(self, request):
        """Statistiques des heures supplémentaires"""
        from django.db.models import Sum, Avg, Count
        from datetime import datetime, timedelta
        
        employee_id = request.query_params.get('employee', None)
        start_date = request.query_params.get('start_date', None)
        end_date = request.query_params.get('end_date', None)
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date() - timedelta(days=30)
        
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = timezone.now().date()
        
        queryset = PresenceTracking.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            overtime_hours__gt=0
        )
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        stats = queryset.aggregate(
            total_overtime=Sum('overtime_hours'),
            avg_overtime=Avg('overtime_hours'),
            total_days=Count('id')
        )
        
        # Par employé
        employee_stats = []
        for employee in Employee.objects.filter(id__in=queryset.values_list('employee_id', flat=True).distinct()):
            emp_overtime = queryset.filter(employee=employee).aggregate(
                total=Sum('overtime_hours'),
                days=Count('id')
            )
            if emp_overtime['total']:
                employee_stats.append({
                    'employee': {
                        'id': employee.id,
                        'name': employee.get_full_name(),
                        'employee_id': employee.employee_id
                    },
                    'total_overtime': float(emp_overtime['total']),
                    'days_with_overtime': emp_overtime['days']
                })
        
        return Response({
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_overtime_hours': float(stats['total_overtime'] or 0),
                'average_overtime_per_day': float(stats['avg_overtime'] or 0),
                'total_days_with_overtime': stats['total_days'] or 0
            },
            'by_employee': sorted(employee_stats, key=lambda x: x['total_overtime'], reverse=True)
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporter les pointages en Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        queryset = self.get_queryset()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Pointages"
        
        # En-têtes
        headers = ['Employé', 'ID Employé', 'Date', 'Heure Arrivée', 'Heure Départ', 'Statut', 'Retard (min)', 'Notes']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for tracking in queryset:
            ws.append([
                f"{tracking.employee.first_name} {tracking.employee.last_name}",
                tracking.employee.employee_id,
                tracking.date.strftime('%Y-%m-%d'),
                tracking.check_in_time.strftime('%H:%M:%S') if tracking.check_in_time else '',
                tracking.check_out_time.strftime('%H:%M:%S') if tracking.check_out_time else '',
                tracking.get_status_display(),
                tracking.late_minutes if tracking.is_late else 0,
                tracking.notes
            ])
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"pointages_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Exporter les pointages en PDF"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        
        queryset = self.get_queryset()
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20,
        )
        title = Paragraph("RAPPORT DE POINTAGE", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Données du tableau
        data = [['Employé', 'ID', 'Date', 'Arrivée', 'Départ', 'Statut', 'Retard']]
        
        for tracking in queryset:
            data.append([
                f"{tracking.employee.first_name} {tracking.employee.last_name}",
                tracking.employee.employee_id,
                tracking.date.strftime('%d/%m/%Y'),
                tracking.check_in_time.strftime('%H:%M') if tracking.check_in_time else '-',
                tracking.check_out_time.strftime('%H:%M') if tracking.check_out_time else '-',
                tracking.get_status_display(),
                f"{tracking.late_minutes} min" if tracking.is_late else '-'
            ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"pointages_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class TrainingPlanViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les plans de formation"""
    queryset = TrainingPlan.objects.all()
    serializer_class = TrainingPlanSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('employee', 'employee__service', 'created_by')
        employee_id = self.request.query_params.get('employee', None)
        status_filter = self.request.query_params.get('status', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activer un plan de formation"""
        plan = self.get_object()
        plan.status = 'ACTIVE'
        plan.save()
        return Response({'message': 'Plan de formation activé', 'plan': TrainingPlanSerializer(plan).data})
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Marquer un plan comme terminé"""
        plan = self.get_object()
        plan.status = 'COMPLETED'
        if not plan.end_date:
            from datetime import date
            plan.end_date = date.today()
        plan.save()
        return Response({'message': 'Plan de formation terminé', 'plan': TrainingPlanSerializer(plan).data})
    
    @action(detail=True, methods=['get'])
    def progress(self, request, pk=None):
        """Récupérer le progrès d'un plan de formation"""
        plan = self.get_object()
        trainings = Training.objects.filter(training_plan=plan)
        
        return Response({
            'plan': TrainingPlanSerializer(plan).data,
            'total_trainings': trainings.count(),
            'completed_trainings': trainings.filter(status='COMPLETED').count(),
            'in_progress_trainings': trainings.filter(status='IN_PROGRESS').count(),
            'completion_rate': plan.completion_rate
        })


class TrainingViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les formations"""
    queryset = Training.objects.all()
    serializer_class = TrainingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('employee', 'training_plan', 'created_by')
        employee_id = self.request.query_params.get('employee', None)
        training_plan_id = self.request.query_params.get('training_plan', None)
        status_filter = self.request.query_params.get('status', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if training_plan_id:
            queryset = queryset.filter(training_plan_id=training_plan_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Démarrer une formation"""
        training = self.get_object()
        training.status = 'IN_PROGRESS'
        training.save()
        return Response({'message': 'Formation démarrée', 'training': TrainingSerializer(training).data})
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Marquer une formation comme terminée"""
        training = self.get_object()
        training.status = 'COMPLETED'
        if not training.end_date:
            from datetime import date
            training.end_date = date.today()
        training.save()
        return Response({'message': 'Formation terminée', 'training': TrainingSerializer(training).data})


class TrainingSessionViewSet(viewsets.ModelViewSet):
    """ViewSet pour gérer les sessions de formation"""
    queryset = TrainingSession.objects.all()
    serializer_class = TrainingSessionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('training', 'training__employee')
        training_id = self.request.query_params.get('training', None)
        status_filter = self.request.query_params.get('status', None)
        
        if training_id:
            queryset = queryset.filter(training_id=training_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def mark_attendance(self, request, pk=None):
        """Marquer la présence à une session"""
        session = self.get_object()
        session.attendance = True
        session.status = 'COMPLETED'
        session.save()
        return Response({'message': 'Présence enregistrée', 'session': TrainingSessionSerializer(session).data})


class EvaluationViewSet(viewsets.ModelViewSet):
    queryset = Evaluation.objects.all()
    serializer_class = EvaluationSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('employee', 'evaluated_by', 'approved_by')
        employee_id = self.request.query_params.get('employee', None)
        evaluation_type = self.request.query_params.get('evaluation_type', None)
        status_filter = self.request.query_params.get('status', None)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if evaluation_type:
            queryset = queryset.filter(evaluation_type=evaluation_type)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.order_by('-evaluation_date', '-created_at')
    
    def perform_create(self, serializer):
        serializer.save(evaluated_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def add_manager_feedback(self, request, pk=None):
        """Ajouter le feedback du manager"""
        evaluation = self.get_object()
        
        manager_feedback = request.data.get('manager_feedback', '')
        manager_recommendations = request.data.get('manager_recommendations', '')
        
        evaluation.manager_feedback = manager_feedback
        evaluation.manager_recommendations = manager_recommendations
        evaluation.status = 'IN_PROGRESS'
        evaluation.save()
        
        return Response({
            'message': 'Feedback du manager ajouté',
            'evaluation': EvaluationSerializer(evaluation).data
        })
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approuver une évaluation"""
        evaluation = self.get_object()
        
        if evaluation.status != 'COMPLETED':
            return Response(
                {'error': 'L\'évaluation doit être terminée avant d\'être approuvée'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        evaluation.status = 'APPROVED'
        evaluation.approved_by = request.user
        evaluation.approval_date = timezone.now()
        evaluation.save()
        
        return Response({
            'message': 'Évaluation approuvée',
            'evaluation': EvaluationSerializer(evaluation).data
        })
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Marquer une évaluation comme terminée"""
        evaluation = self.get_object()
        evaluation.status = 'COMPLETED'
        evaluation.save()
        
        return Response({
            'message': 'Évaluation terminée',
            'evaluation': EvaluationSerializer(evaluation).data
        })
    
    @action(detail=False, methods=['get'])
    def pending_feedback(self, request):
        """Récupérer les évaluations en attente de feedback manager"""
        pending = Evaluation.objects.filter(
            status__in=['DRAFT', 'IN_PROGRESS'],
            manager_feedback=''
        ).select_related('employee', 'evaluated_by')
        
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def annual(self, request):
        """Récupérer les évaluations annuelles"""
        from datetime import datetime
        year = request.query_params.get('year', datetime.now().year)
        
        annual_evaluations = Evaluation.objects.filter(
            evaluation_type='ANNUAL',
            evaluation_date__year=year
        ).select_related('employee', 'evaluated_by', 'approved_by')
        
        serializer = self.get_serializer(annual_evaluations, many=True)
        return Response({
            'year': year,
            'total': annual_evaluations.count(),
            'evaluations': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporter les évaluations en Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.http import HttpResponse
        except ImportError:
            return Response(
                {'error': 'openpyxl n\'est pas installé. Installez-le avec: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        queryset = self.get_queryset()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Évaluations"
        
        # En-têtes
        headers = ['Employé', 'ID Employé', 'Date', 'Type', 'Note (sur 5)', 'Commentaires', 'Évalué par']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for evaluation in queryset:
            ws.append([
                f"{evaluation.employee.first_name} {evaluation.employee.last_name}",
                evaluation.employee.employee_id,
                evaluation.evaluation_date.strftime('%Y-%m-%d'),
                evaluation.get_evaluation_type_display(),
                float(evaluation.performance_score),
                evaluation.comments,
                evaluation.evaluated_by.get_full_name() if evaluation.evaluated_by else 'N/A'
            ])
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"evaluations_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Exporter les évaluations en PDF"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        
        queryset = self.get_queryset()
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=20,
        )
        title = Paragraph("RAPPORT D'ÉVALUATIONS", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Données du tableau
        data = [['Employé', 'ID', 'Date', 'Type', 'Note/5', 'Commentaires']]
        
        for evaluation in queryset:
            comments = evaluation.comments[:50] + '...' if len(evaluation.comments) > 50 else evaluation.comments
            data.append([
                f"{evaluation.employee.first_name} {evaluation.employee.last_name}",
                evaluation.employee.employee_id,
                evaluation.evaluation_date.strftime('%d/%m/%Y'),
                evaluation.get_evaluation_type_display(),
                str(float(evaluation.performance_score)),
                comments or '-'
            ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch, 0.8*inch, 3.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Centrer la colonne Note
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f"evaluations_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response